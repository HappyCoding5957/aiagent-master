"""
ESG Upload 系統 - FastAPI 版本（統一架構）
端口：8001
功能：附件三知識庫上傳管理，整合統一 SSO
路由：/esg/upload/
"""
from fastapi import FastAPI, Request, Depends, HTTPException, File, UploadFile
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import os
import json
import logging
import tempfile
import hashlib
import threading
import requests
import psycopg2
import openpyxl
from pathlib import Path
from datetime import datetime
from typing import Optional
from uuid import uuid4
from pydantic import BaseModel

# 統一 SSO 中間件
from simple_auth import get_current_user, require_login

# 配置日誌
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ========== 配置 ==========
DB_HOST = "10.100.40.5"
DB_PORT = 8002
DB_USER = "dgtk"
DB_PASS = "dgtk"
DB_NAME = "dgtk"
EMBED_API = "http://10.100.40.5:8004/api/embed"
UNIT = "SYSTEM"
PDF_NAME = "附件三_EnvSafety_atta3_知識庫"

BASE_DIR = Path(__file__).parent
PROGRESS_FILE = str(BASE_DIR / "upload_progress.json")
UPLOAD_LOG_FILE = str(BASE_DIR / "upload_logs.jsonl")
STATIC_DIR = BASE_DIR / "static"
TEMPLATES_DIR = BASE_DIR / "templates"
UPLOADS_DIR = BASE_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

# ========== 全局上傳狀態 ==========
upload_status = {"is_uploading": False}

# ========== 創建 FastAPI app ==========
app = FastAPI(title="ESG Upload System", version="2.0")
app.mount("/esg/upload/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# ========== Pydantic Models ==========
class DeleteResponse(BaseModel):
    success: bool
    deleted_chunks: int
    deleted_files: int
    message: str

class ProgressResponse(BaseModel):
    stage: str
    percent: int
    message: str
    timestamp: Optional[str] = None
    is_uploading: Optional[bool] = False


# ========== DB 連線 ==========
def log_upload(user: str, user_name: str, file_name: str, status: str, chunks: int = 0, details: str = ""):
    """記錄上傳日誌至 PostgreSQL"""
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute(
            """INSERT INTO esg_usage_logs (timestamp, service, user_id, user_name, file_name, status, chunks, details)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
            (datetime.now(), 'upload', user, user_name, file_name, status, chunks, details)
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        logger.warning(f"[LOG] 無法寫入 DB 日誌: {e}")


def get_db_conn():
    return psycopg2.connect(
        host=DB_HOST, port=DB_PORT,
        user=DB_USER, password=DB_PASS,
        database=DB_NAME
    )


# ========== 進度更新 ==========
def update_progress(stage: str, percent: int, message: str = ""):
    progress = {
        "stage": stage,
        "percent": percent,
        "message": message,
        "timestamp": datetime.now().isoformat()
    }
    try:
        with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
            json.dump(progress, f, ensure_ascii=False)
    except Exception as e:
        logger.warning(f"[PROGRESS] 無法寫入進度: {e}")
    logger.info(f"[PROGRESS] {stage} - {percent}% - {message}")


# ========== 後台上傳處理 ==========
def process_upload(file_path: str, user_id: str = "unknown", user_name: str = "未知", original_filename: str = ""):
    try:
        update_progress("init", 0, "開始處理")

        # 1. 讀取 Excel
        update_progress("reading", 5, "讀取 Excel 檔案")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        rows_data = []
        for row_idx in range(2, ws.max_row + 1):
            cols = [ws.cell(row_idx, c).value or "" for c in range(1, 8)]
            if not any(cols[:6]):
                continue
            chunk_text = (
                f"[類別] {cols[0]}\n[行為準則] {cols[1]}\n[關鍵字] {cols[2]}\n"
                f"[條文內容] {cols[3]}\n[權責部門] {cols[4]}\n[目前現況] {cols[5]}\n[問卷出處] {cols[6]}"
            )
            rows_data.append({"row_index": row_idx, "text": chunk_text})

        wb.close()
        update_progress("reading", 10, f"讀取完成，共 {len(rows_data)} 筆")

        if not rows_data:
            raise ValueError("沒有資料可上傳")

        # 2. 生成 Embedding
        update_progress("embedding", 15, "開始生成向量")
        texts = [r["text"] for r in rows_data]
        vectors = []
        batch_size = 50
        total_batches = (len(texts) + batch_size - 1) // batch_size

        for i in range(0, len(texts), batch_size):
            batch_texts = texts[i:i + batch_size]
            batch_num = i // batch_size + 1
            progress_pct = 15 + int((batch_num / total_batches) * 45)
            update_progress("embedding", progress_pct, f"生成向量 {batch_num}/{total_batches} 批")

            resp = requests.post(EMBED_API, json={"input": batch_texts}, timeout=120)
            resp.raise_for_status()
            result = resp.json()
            if "vectors" in result:
                vectors.extend(result["vectors"])
            elif "data" in result:
                vectors.extend([item["embedding"] for item in result["data"]])
            else:
                raise ValueError(f"無法解析 embedding 格式: {list(result.keys())}")

        update_progress("embedding", 60, "向量生成完成")

        # 3. 寫入資料庫
        update_progress("database", 65, "開始寫入資料庫")
        conn = get_db_conn()
        cur = conn.cursor()

        # 刪除舊資料
        cur.execute(
            "SELECT id FROM pdffile WHERE name = %s AND unit = %s",
            (PDF_NAME, UNIT)
        )
        old = cur.fetchone()
        if old:
            update_progress("database", 70, "刪除舊資料")
            cur.execute("DELETE FROM pdfchunk WHERE pdf_id = %s", (old[0],))
            cur.execute("DELETE FROM pdffile WHERE id = %s", (old[0],))
            conn.commit()

        # 建立新的 pdffile
        update_progress("database", 75, "建立檔案記錄")
        pdf_file_id = str(uuid4())
        pdf_hash = hashlib.md5(PDF_NAME.encode()).hexdigest()
        cur.execute(
            "INSERT INTO pdffile (id, hash, unit, name, size, date) VALUES (%s, %s, %s, %s, %s, %s)",
            (pdf_file_id, pdf_hash, UNIT, PDF_NAME, str(len(rows_data)), datetime.now())
        )
        conn.commit()

        # 寫入 pdfchunk
        update_progress("database", 80, "寫入資料 chunks")
        for idx, (row_data, vector) in enumerate(zip(rows_data, vectors)):
            cur.execute(
                "INSERT INTO pdfchunk (id, pdf_id, page_hash, chunk_index, xywh, text, embed) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s)",
                (
                    str(uuid4()), pdf_file_id,
                    f"row_{row_data['row_index']}", idx, "",
                    row_data["text"], json.dumps(vector)
                )
            )

        update_progress("database", 90, "提交資料庫")
        conn.commit()
        cur.close()
        conn.close()

        update_progress("complete", 100, "上傳完成")
        logger.info(f"[UPLOAD] 完成，共 {len(rows_data)} 筆")
        # 記錄日誌
        log_upload(user_id, user_name, original_filename or PDF_NAME, "success", len(rows_data))

    except Exception as e:
        update_progress("error", 0, f"錯誤: {str(e)}")
        logger.error(f"[UPLOAD] 失敗: {e}")
    finally:
        upload_status["is_uploading"] = False
        try:
            os.unlink(file_path)
        except Exception:
            pass


# ========== Routes ==========

@app.get("/esg/upload/favicon.ico")
async def favicon():
    favicon_path = STATIC_DIR / "favicon.ico"
    if favicon_path.exists():
        return FileResponse(favicon_path)
    return JSONResponse({"error": "Not Found"}, status_code=404)


@app.get("/esg/upload/", response_class=HTMLResponse)
@app.get("/esg/upload", response_class=HTMLResponse)
async def upload_page(request: Request, user: dict = Depends(require_login)):
    logger.info(f"[SSO] 用戶 {user['family_name']} ({user['preferred_username']}) 存取上傳頁面")
    html_file = TEMPLATES_DIR / "upload.html"
    if html_file.exists():
        with open(html_file, "r", encoding="utf-8") as f:
            html_content = f.read()
        html_content = html_content.replace("{{user_name}}", user['family_name'])
        html_content = html_content.replace("{{user_id}}", user['preferred_username'])
        html_content = html_content.replace("{{user_dep}}", user.get('dep', ''))
        return HTMLResponse(content=html_content)
    return HTMLResponse("<h1>ESG Upload System</h1><p>模板文件不存在</p>")


@app.post("/esg/upload/api/upload")
async def api_upload(
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(require_login)
):
    """非同步上傳：立即返回，後台處理 Excel→Embedding→DB"""
    global upload_status

    # 檢查是否已有上傳進行中
    if os.path.exists(PROGRESS_FILE):
        try:
            with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
                prog = json.load(f)
            if prog.get("stage") not in ["idle", "complete", "error"]:
                raise HTTPException(
                    status_code=409,
                    detail=f"已有上傳進行中（{prog.get('percent')}%），請稍後再試"
                )
        except json.JSONDecodeError:
            pass

    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="只接受 .xlsx 檔案")

    # 儲存暫存檔案
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    update_progress("init", 0, "開始處理")
    upload_status["is_uploading"] = True

    # 啟動後台執行緒
    thread = threading.Thread(target=process_upload, args=(tmp_path, user["preferred_username"], user.get("family_name", "未知"), file.filename), daemon=True)
    thread.start()

    logger.info(f"[UPLOAD] {user['preferred_username']} 啟動後台處理: {file.filename}")

    return JSONResponse({
        "status": "success",
        "message": "上傳已開始，請輪詢 /api/progress 查看進度"
    })


@app.delete("/esg/upload/api/delete")
async def api_delete(request: Request, user: dict = Depends(require_login)):
    """刪除附件三知識庫"""
    try:
        conn = get_db_conn()
        cur = conn.cursor()

        cur.execute(
            "SELECT id FROM pdffile WHERE name = %s AND unit = %s",
            (PDF_NAME, UNIT)
        )
        row = cur.fetchone()

        if not row:
            cur.close()
            conn.close()
            return JSONResponse({
                "success": True,
                "deleted_chunks": 0,
                "deleted_files": 0,
                "message": "知識庫不存在"
            })

        pdf_id = row[0]
        cur.execute("SELECT COUNT(*) FROM pdfchunk WHERE pdf_id = %s", (pdf_id,))
        deleted_chunks = cur.fetchone()[0]

        cur.execute("DELETE FROM pdfchunk WHERE pdf_id = %s", (pdf_id,))
        cur.execute("DELETE FROM pdffile WHERE id = %s", (pdf_id,))
        conn.commit()
        cur.close()
        conn.close()

        logger.info(f"[DELETE] {user['preferred_username']} 刪除知識庫，chunks: {deleted_chunks}")

        return JSONResponse({
            "success": True,
            "deleted_chunks": deleted_chunks,
            "deleted_files": 1,
            "message": "刪除成功"
        })

    except Exception as e:
        logger.error(f"[DELETE] 失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/esg/upload/api/progress")
async def api_progress(request: Request):
    """查詢上傳進度（每 2 秒輪詢）"""
    try:
        if os.path.exists(PROGRESS_FILE):
            with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
                progress = json.load(f)
            progress["is_uploading"] = upload_status.get("is_uploading", False)
            return JSONResponse(progress)
        else:
            return JSONResponse({
                "stage": "idle",
                "percent": 0,
                "message": "尚未開始",
                "is_uploading": False
            })
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/esg/upload/api/files")
async def api_files(request: Request, user: dict = Depends(require_login)):
    """獲取已上傳檔案列表"""
    try:
        files = []
        for file_path in UPLOADS_DIR.glob("*"):
            if file_path.is_file():
                stat = file_path.stat()
                files.append({
                    "filename": file_path.name,
                    "size": stat.st_size,
                    "modified": datetime.fromtimestamp(stat.st_mtime).isoformat()
                })
        return JSONResponse({"status": "success", "files": files, "total": len(files)})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/esg/upload/api/health")
async def health_check(request: Request):
    """健康檢查 + 知識庫狀態"""
    result = {
        "status": "healthy",
        "service": "ESG Upload System (FastAPI)",
        "version": "2.0",
        "port": 8001
    }
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT id, name, date FROM pdffile WHERE unit = %s AND name = %s ORDER BY date DESC LIMIT 1",
            (UNIT, PDF_NAME)
        )
        row = cur.fetchone()
        if row:
            pdf_id, name, date = row
            cur.execute("SELECT COUNT(*) FROM pdfchunk WHERE pdf_id = %s", (pdf_id,))
            chunk_count = cur.fetchone()[0]
            result["kb_status"] = "exists"
            result["kb_name"] = name
            result["kb_chunks"] = chunk_count
            result["kb_updated"] = date.strftime("%Y-%m-%d %H:%M:%S") if date else "-"
            result["kb_pdf_id"] = str(pdf_id)
        else:
            result["kb_status"] = "not_exists"
            result["kb_name"] = "-"
            result["kb_chunks"] = 0
            result["kb_updated"] = "-"
            result["kb_pdf_id"] = "-"
        cur.close()
        conn.close()
    except Exception as e:
        logger.error(f"[Health] 失敗: {e}")
        result["kb_status"] = "error"
    return JSONResponse(result)


@app.get("/esg/upload/api/status")
async def api_status(request: Request):
    """狀態檢查：返回當前用戶登入狀態"""
    user = get_current_user(request)
    if user:
        return JSONResponse({"logged_in": True, "user": user})
    return JSONResponse({"logged_in": False}, status_code=401)


@app.get("/esg/upload/api/stats")
async def get_upload_stats(request: Request, user: dict = Depends(require_login)):
    """查詢上傳統計資訊(從 PostgreSQL)"""
    try:
        conn = get_db_conn()
        cur = conn.cursor()

        today = datetime.now().date()

        cur.execute("SELECT COUNT(*) FROM esg_usage_logs WHERE service='upload'")
        total = cur.fetchone()[0]

        cur.execute("SELECT COUNT(*) FROM esg_usage_logs WHERE service='upload' AND timestamp::date = %s", (today,))
        today_count = cur.fetchone()[0]

        cur.execute("SELECT COUNT(DISTINCT user_id) FROM esg_usage_logs WHERE service='upload'")
        total_users = cur.fetchone()[0]

        cur.execute("""SELECT timestamp, user_id, user_name, file_name, status, chunks
                        FROM esg_usage_logs WHERE service='upload'
                        ORDER BY timestamp DESC LIMIT 200""")
        rows = cur.fetchall()
        cur.close()
        conn.close()

        logs = [
            {"timestamp": r[0].isoformat(), "user": r[1], "user_name": r[2],
             "file_name": r[3], "status": r[4], "chunks": r[5], "action": "upload"}
            for r in rows
        ]
        return {"total_requests": total, "today_requests": today_count,
                "total_users": total_users, "logs": logs}
    except Exception as e:
        logger.error(f"[STATS] DB 查詢失敗: {e}")
        return {"total_requests": 0, "today_requests": 0, "total_users": 0, "logs": []}


@app.get("/esg/upload/api/logout")
async def api_logout(request: Request):
    """登出"""
    from simple_auth import logout
    return await logout(request)


@app.get("/esg/upload/admin", response_class=HTMLResponse)
async def admin_page(request: Request, user: dict = Depends(require_login)):
    """管理員統計頁面"""
    html_file = TEMPLATES_DIR / "admin.html"
    if html_file.exists():
        with open(html_file, "r", encoding="utf-8") as f:
            html_content = f.read()
        html_content = html_content.replace("{{user_name}}", user['family_name'])
        html_content = html_content.replace("{{user_id}}", user['preferred_username'])
        return HTMLResponse(content=html_content, headers={"Cache-Control": "no-cache, no-store, must-revalidate"})
    return HTMLResponse("<h1>Admin</h1><p>模板文件不存在</p>")


if __name__ == "__main__":
    import uvicorn
    print("=" * 60)
    print("🚀 ESG Upload System (FastAPI) 啟動")
    print("Port: 8001  |  /esg/upload/")
    print("APIs: upload / delete / progress / health / status / files")
    print("=" * 60)
    uvicorn.run(app, host="0.0.0.0", port=8001, log_level="info")
