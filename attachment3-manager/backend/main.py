#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FastAPI 後端 - 職安問卷上傳系統
提供上傳、查詢、刪除 API

修改：改為異步處理，支援實時進度反饋
"""
import os
import json
import base64
import tempfile
import hashlib
import threading
from datetime import datetime
from pathlib import Path
from typing import Optional
from uuid import uuid4

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from sqlmodel import Session, create_engine, select
import openpyxl
import requests

# 導入資料庫模型
from models import PdfFile, PdfChunk, SQLModel

# ========== 配置 ==========
DB_URL = os.getenv("DB_URL", "postgresql://dgtk:dgtk@10.100.40.5:8002/dgtk")
EMBED_API = os.getenv("EMBED_API", "http://10.100.40.5:8004/api/embed")
UNIT = "SYSTEM"
PDF_NAME = "附件三_EnvSafety_atta3_知識庫"
PROGRESS_FILE = "/app/attachment3_upload_progress.json"

# ========== 全局狀態追蹤 ==========
upload_status = {
    "is_uploading": False,
    "start_time": None
}

# ========== FastAPI 應用 ==========
app = FastAPI(title="職安問卷上傳系統", version="1.0.0")

# CORS 設定（允許前端跨域請求）
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 生產環境建議限制特定域名
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 建立資料庫引擎
engine = create_engine(DB_URL, echo=False)


# ========== Pydantic Models ==========
class StatusResponse(BaseModel):
    exists: bool
    pdf_id: Optional[str] = None
    name: Optional[str] = None
    chunk_count: Optional[int] = None
    last_update: Optional[str] = None
    unit: Optional[str] = None


class UploadResponse(BaseModel):
    success: bool
    message: str
    start_time: Optional[str] = None
    pdf_id: Optional[str] = None
    pdf_name: Optional[str] = None
    chunk_count: Optional[int] = None
    unit: Optional[str] = None
    error: Optional[str] = None


class DeleteResponse(BaseModel):
    success: bool
    deleted_chunks: int
    deleted_files: int
    message: str


class ProgressResponse(BaseModel):
    stage: str
    percent: int
    message: str
    timestamp: Optional[str] = None
    is_uploading: Optional[bool] = False  # ✅ 新增欄位


# ========== 輔助函數 ==========
def update_progress(stage: str, percent: int, message: str = ""):
    """更新進度到檔案"""
    progress = {
        "stage": stage,
        "percent": percent,
        "message": message,
        "timestamp": datetime.now().isoformat()
    }
    try:
        with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
            json.dump(progress, f, ensure_ascii=False)
        print(f"[PROGRESS] {stage} - {percent}% - {message}")  # ✅ 添加日誌
    except Exception as e:
        print(f"⚠️  無法寫入進度檔案: {e}")


def process_upload(file_path: str) -> dict:
    """處理上傳的附件三檔案"""
    try:
        update_progress("init", 0, "開始處理")

        # ========== 1. 讀取 Excel ==========
        update_progress("reading", 5, "讀取 Excel 檔案")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        rows_data = []
        for row_idx in range(2, ws.max_row + 1):
            a_col = ws[f"A{row_idx}"].value or ""
            b_col = ws[f"B{row_idx}"].value or ""
            c_col = ws[f"C{row_idx}"].value or ""
            d_col = ws[f"D{row_idx}"].value or ""
            e_col = ws[f"E{row_idx}"].value or ""
            f_col = ws[f"F{row_idx}"].value or ""
            g_col = ws[f"G{row_idx}"].value or ""

            if not any([a_col, b_col, c_col, d_col, e_col, f_col]):
                continue

            chunk_text = f"""[類別] {a_col}
[行為準則] {b_col}
[關鍵字] {c_col}
[條文內容] {d_col}
[權責部門] {e_col}
[目前現況] {f_col}
[問卷出處] {g_col}"""

            rows_data.append({
                "row_index": row_idx,
                "text": chunk_text
            })

        wb.close()
        update_progress("reading", 10, f"讀取完成，共 {len(rows_data)} 筆")

        if not rows_data:
            raise ValueError("沒有資料可上傳")

        # ========== 2. 生成 Embedding ==========
        update_progress("embedding", 15, "開始生成向量")
        texts = [row["text"] for row in rows_data]
        vectors = []
        batch_size = 50

        total_batches = (len(texts) + batch_size - 1) // batch_size

        for i in range(0, len(texts), batch_size):
            batch_texts = texts[i:i + batch_size]
            batch_num = i // batch_size + 1

            # ✅ 計算進度 (15% -> 60%)
            progress_percent = 15 + int((batch_num / total_batches) * 45)
            update_progress("embedding", progress_percent, f"生成向量 {batch_num}/{total_batches} 批")

            resp = requests.post(
                EMBED_API,
                json={"input": batch_texts},
                timeout=120
            )
            resp.raise_for_status()
            result = resp.json()

            # 解析 vectors
            if "vectors" in result:
                batch_vectors = result["vectors"]
            elif "data" in result:
                batch_vectors = [item["embedding"] for item in result["data"]]
            else:
                raise ValueError(f"無法解析 embedding 回傳格式: {result.keys()}")

            vectors.extend(batch_vectors)

        update_progress("embedding", 60, "向量生成完成")

        # ========== 3. 插入資料庫 ==========
        update_progress("database", 65, "開始寫入資料庫")

        with Session(engine) as db:
            # 檢查並刪除舊資料
            existing = db.exec(
                select(PdfFile).where(
                    PdfFile.name == PDF_NAME,
                    PdfFile.unit == UNIT
                )
            ).first()

            if existing:
                update_progress("database", 70, "刪除舊資料")
                # 刪除舊的 chunks
                for chunk in db.exec(select(PdfChunk).where(PdfChunk.pdf_id == existing.id)):
                    db.delete(chunk)
                db.delete(existing)
                db.commit()

            # 建立新的 PdfFile
            update_progress("database", 75, "建立檔案記錄")
            pdf_file_id = str(uuid4())
            pdf_hash = hashlib.md5(PDF_NAME.encode()).hexdigest()

            pdf_file = PdfFile(
                id=pdf_file_id,
                hash=pdf_hash,
                unit=UNIT,
                name=PDF_NAME,
                size=str(len(rows_data)),
                date=datetime.now()
            )
            db.add(pdf_file)
            db.commit()

            # 建立 PdfChunk
            update_progress("database", 80, "寫入資料 chunks")
            for idx, (row_data, vector) in enumerate(zip(rows_data, vectors)):
                chunk = PdfChunk(
                    id=str(uuid4()),
                    pdf_id=pdf_file_id,
                    page_hash=f"row_{row_data['row_index']}",
                    chunk_index=idx,
                    xywh="",
                    text=row_data["text"],
                    embed=vector
                )
                db.add(chunk)

            update_progress("database", 90, "提交資料庫")
            db.commit()

        update_progress("complete", 100, "上傳完成")

        return {
            "success": True,
            "pdf_id": pdf_file_id,
            "pdf_name": PDF_NAME,
            "chunk_count": len(rows_data),
            "unit": UNIT
        }

    except Exception as e:
        update_progress("error", 0, f"錯誤: {str(e)}")
        return {
            "success": False,
            "error": str(e)
        }


# ========== API 端點 ==========
@app.get("/")
def read_root():
    """根路徑"""
    return {"message": "職安問卷上傳系統 API", "version": "1.0.1 (async)"}


@app.get("/api/status", response_model=StatusResponse)
def get_status():
    """查詢知識庫狀態"""
    try:
        with Session(engine) as db:
            pdf_file = db.exec(
                select(PdfFile).where(
                    PdfFile.name == PDF_NAME,
                    PdfFile.unit == UNIT
                )
            ).first()

            if pdf_file:
                chunk_count = db.exec(
                    select(PdfChunk).where(PdfChunk.pdf_id == pdf_file.id)
                ).all()

                return StatusResponse(
                    exists=True,
                    pdf_id=pdf_file.id,
                    name=pdf_file.name,
                    chunk_count=len(chunk_count),
                    last_update=pdf_file.date.isoformat() if pdf_file.date else None,
                    unit=pdf_file.unit
                )
            else:
                return StatusResponse(exists=False)

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/upload", response_model=UploadResponse)
async def upload_file(file: UploadFile = File(...)):
    """
    ✅ 異步上傳 - 立即返回，後台處理
    前端應該在收到響應後立即開始輪詢 /api/progress
    """
    global upload_status

    try:
        # ✅ 檢查是否已有上傳在進行（基於進度檔案，更穩健）
        if os.path.exists(PROGRESS_FILE):
            try:
                with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
                    progress = json.load(f)
                    # 如果進度不是 idle, complete, error，表示正在處理
                    if progress.get("stage") not in ["idle", "complete", "error"]:
                        raise HTTPException(
                            status_code=409,
                            detail=f"已有上傳正在進行中（{progress.get('percent')}%），請稍後再試"
                        )
            except json.JSONDecodeError:
                pass  # 檔案損壞，允許繼續

        # 檢查檔案類型
        if not file.filename.endswith('.xlsx'):
            raise HTTPException(status_code=400, detail="只接受 .xlsx 檔案")

        # 儲存暫存檔案
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_file_path = tmp_file.name

        # ✅ 初始化進度（讓前端立即看到 0%）
        update_progress("init", 0, "開始處理")

        # 標記上傳中
        upload_status["is_uploading"] = True
        upload_status["start_time"] = datetime.now().isoformat()

        # ✅ 啟動後台執行緒處理
        def background_process():
            try:
                print(f"[UPLOAD] 開始後台處理: {file.filename}")
                result = process_upload(tmp_file_path)
                print(f"[UPLOAD] 處理完成: {result}")

                # 清理暫存檔案
                try:
                    os.unlink(tmp_file_path)
                except Exception as cleanup_error:
                    print(f"[UPLOAD] 清理暫存檔案失敗: {cleanup_error}")

            except Exception as e:
                print(f"[UPLOAD] 處理失敗: {e}")
                update_progress("error", 0, f"上傳失敗: {str(e)}")
            finally:
                upload_status["is_uploading"] = False
                print("[UPLOAD] 後台處理結束")

        thread = threading.Thread(target=background_process, daemon=True)
        thread.start()

        print(f"[UPLOAD] 已啟動後台處理，立即返回響應")

        # ✅ 立即返回（不等待處理完成）
        return UploadResponse(
            success=True,
            message="上傳已開始，請輪詢 /api/progress 查看進度（建議每 2 秒輪詢一次）",
            start_time=upload_status["start_time"],
            pdf_id="PROCESSING",
            pdf_name="處理中",
            chunk_count=0,
            unit="SYSTEM"
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/delete", response_model=DeleteResponse)
def delete_knowledge_base():
    """刪除知識庫"""
    try:
        with Session(engine) as db:
            # 查詢檔案
            pdf_file = db.exec(
                select(PdfFile).where(
                    PdfFile.name == PDF_NAME,
                    PdfFile.unit == UNIT
                )
            ).first()

            if not pdf_file:
                return DeleteResponse(
                    success=True,
                    deleted_chunks=0,
                    deleted_files=0,
                    message="知識庫不存在"
                )

            # 刪除 chunks
            chunks = db.exec(select(PdfChunk).where(PdfChunk.pdf_id == pdf_file.id)).all()
            deleted_chunks = len(chunks)
            for chunk in chunks:
                db.delete(chunk)

            # 刪除 file
            db.delete(pdf_file)
            db.commit()

            return DeleteResponse(
                success=True,
                deleted_chunks=deleted_chunks,
                deleted_files=1,
                message="刪除成功"
            )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/progress", response_model=ProgressResponse)
def get_progress():
    """
    ✅ 查詢上傳進度
    前端應該每 2 秒輪詢一次此 API
    """
    try:
        if os.path.exists(PROGRESS_FILE):
            with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
                progress = json.load(f)

                # ✅ 添加上傳狀態
                progress["is_uploading"] = upload_status.get("is_uploading", False)

                return ProgressResponse(**progress)
        else:
            return ProgressResponse(
                stage="idle",
                percent=0,
                message="尚未開始",
                is_uploading=upload_status.get("is_uploading", False)
            )
    except Exception as e:
        print(f"[PROGRESS] 讀取進度失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
