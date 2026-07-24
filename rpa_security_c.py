#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
客戶問卷 RPA - 核心比對邏輯
功能：
  1. 讀取附件2(問卷)和附件3(資料庫)
  2. 執行兩階段比對(行為準則 → 關鍵字)
  3. 回寫 D 欄(權責部門)和 E 欄(現況/影響)
  4. 產生比對報告

修改版本：v2 - 修復 Excel 自動換行問題 + 減少修復錯誤
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment  # ★ 新增 Alignment
from rapidfuzz import fuzz
import jieba
import os
from datetime import datetime
import requests  # 用於調用 Azure OpenAI
import re  # ★ 新增 re，做文字清洗用

# ========== LLM-first Schema 萃取 ==========
import sys
import json
sys.path.insert(0, '/home/ifm02web/aiagent')

# ✅ 對齊 upload_attachment3_to_rag.py：優先使用新版函式
try:
    from llm_first_schema import (
        AzureOpenAIConfig,
        extract_question_schema_llm,
        QuestionItem,
    )
    _HAS_LLM_FIRST = True
except Exception as e:
    print(f"⚠️  LLM-first schema 模組載入失敗: {e}")
    _HAS_LLM_FIRST = False

# ✅ 受控 Facet 白名單系統（防跨題域漂移）
try:
    from facet_taxonomy import (
        get_taxonomy_entry,
        domain_gate_filter,
        get_controlled_facets_cached,
        validate_facet_candidate,
        clear_facet_cache,
    )
    _HAS_FACET_TAXONOMY = True
except Exception as e:
    print(f"⚠️  facet_taxonomy 模組載入失敗: {e}")
    _HAS_FACET_TAXONOMY = False


# ========== 比對配置 ==========
MATCH_CONFIG = {
    "phase1_primary_threshold": 60,      # 第一階段主要閾值（模糊匹配用，RAG 不使用）
    "phase1_fallback_threshold": 50,     # 第一階段備援閾值（模糊匹配用，RAG 不使用）
    "keyword_weight": 0.6,               # 關鍵字權重（模糊匹配用，RAG 不使用）
    "behavior_weight": 0.4,              # 行為準則權重（模糊匹配用，RAG 不使用）
    "aggregate_bonus": 5,                # 彙整版加分（模糊匹配用，RAG 不使用）
    "low_confidence_threshold": 55,      # 低信心閾值：< 55 視為紅色（RAG 相似度 0-100%）
    "high_confidence_threshold": 70,     # 高信心閾值：≥ 70 視為綠色（RAG 相似度 0-100%）

    # ── Semantic Score Calibration ──────────────────────────────────────────
    # 換 embedding 模型時，只需更新這 4 個數字，其餘 pipeline 不用動。
    # 設定方法：跑 10~20 題已知正確答案，觀察 raw cosine score 的分布範圍，
    # 將典型正確匹配的 [min, max] 填入 sem_score_min/max。
    #
    # 目前模型 (2026-03 更換)：正確匹配 score 約落在 0.07 ~ 0.15
    # 若 score >= sem_score_max → 直接 * 100（高分不壓縮）
    # 若 sem_score_min <= score < sem_score_max → 線性映射到 [conf_min, conf_max]
    # 若 score < sem_score_min → clamp 到 conf_min（最低保留進黃色審查）
    "sem_score_min": 0.07,   # 新模型：正確匹配 cosine 下限（換模型時更新）
    "sem_score_max": 0.15,   # 新模型：正確匹配 cosine 上限（換模型時更新）
    "sem_conf_min": 55.0,    # sem_score_min 映射到的 confidence%（對應 low_confidence_threshold）
    "sem_conf_max": 95.0,    # sem_score_max 映射到的 confidence%
    # ────────────────────────────────────────────────────────────────────────
}

# ========== Rerank 功能開關 ==========
ENABLE_RERANK = False  # True: 啟用 Rerank 優化 | False: 使用原始 RAG 結果

# ========== Excel 文字清洗 ==========

# Excel 不接受的控制字元（除了 \t, \n 以外的 0x00-0x1F）
_INVALID_CHAR_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

def clean_excel_text(value) -> str:
    """
    將任意輸入轉成適合寫入 Excel 的安全字串：
      - None → ""
      - 統一換行符號
      - 移除 Excel 不喜歡的控制字元（避免開啟時修復）
    """
    if value is None:
        return ""
    s = str(value)
    # 統一換行
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    # 移除控制字元
    s = _INVALID_CHAR_RE.sub("", s)
    return s



# ========== 輔助函數：題型識別與 Rerank 優化 ==========

def is_commitment_statement(question: str) -> bool:
    """
    識別「純粹聲明/知悉」類題目（不含實質 RBA 要求，不需 RAG 比對）

    NOTE: 「承諾」「同意」已移除，因為 DB 有 RBA 承諾類行為準則需要比對。
    只保留純粹儀式性聲明，例如「完全知悉」「供應商確認已閱讀」。
    """
    commitment_keywords = [
        "完全知悉", "知悉並理解", "充分理解",
        "乙方完全", "供應商確認", "接受稽核",
        "同意接受稽核",
    ]
    return any(kw in question for kw in commitment_keywords)


def _normalize_topic_token(token: str) -> str:
    """將主題名稱正規化，供 coverage 比對用（基礎 trim 即可）。"""
    t = (token or "").strip()
    return t


def _parse_facets_from_enum(question: str) -> list:
    """
    從題目括號列舉自動解析面向（Fallback：q_schema 不可用時）。
    支援 (例:A、B、C等) / （例：A、B、C） / (A, B, C) 形式。
    """
    import re as _re
    q = question or ""
    # 比對 (例:...) 或 （例：...）
    m = _re.search(r'[（(][例例如][:：]([^）)]{2,160})[）)]', q)
    if not m:
        # 比對一般括號內含頓號/逗號列舉
        m = _re.search(r'[（(]([^（）()]{4,160}[、，,][^（）()]{2,80})[）)]', q)
    if m:
        inner = m.group(1)
        items = _re.split(r'[、,，/]', inner)
        result = [x.strip().rstrip('等') for x in items if len(x.strip()) >= 2]
        if len(result) >= 2:
            return result
    return []


def get_facets_expected_from_question(question: str, q_schema=None) -> list:
    """
    題目 facets（Expected）優先級：
      1. q_schema.must_have_terms — LLM 每題動態生成，中文錨點詞（零額外呼叫）
      2. _parse_facets_from_enum() — regex 解析題目括號列舉
      3. []                        — 非多面向題
    """
    # 1. LLM-first：q_schema.must_have_terms 即為中文面向錨點
    if q_schema is not None:
        terms = getattr(q_schema, 'must_have_terms', None) or []
        if len(terms) >= 2:
            return [str(t) for t in terms if t]

    # 2. Regex fallback：解析括號列舉
    facets = _parse_facets_from_enum(question)
    if len(facets) >= 2:
        return facets

    return []


def is_multifacet_question_expected(question: str, q_schema=None) -> bool:
    """題目是否多面向：facets >= 2 即視為多面向。"""
    return len(get_facets_expected_from_question(question, q_schema=q_schema)) >= 2


def infer_topic_from_candidate(candidate: dict, vocab: list = None) -> str:
    """
    候選 topics（Observed）。
      vocab = facets_expected（題目錨點詞）→ 動態匹配優先
      vocab = None → 靜態通用 rules 兜底
    """
    matches = candidate.get("matches") or []
    text = " ".join([
        str(candidate.get("behavior") or ""),
        str(candidate.get("article") or ""),
        " ".join([str(m.get("keyword") or "") for m in matches]),
        " ".join([str(m.get("impact") or "") for m in matches]),
    ])

    # 1. Dynamic：加權字元命中數（長詞優先），避免 first-match-wins 偏差
    if vocab:
        best_term, best_score = None, 0
        for term in vocab:
            if not term:
                continue
            cnt = text.count(term)
            if cnt > 0:
                score = cnt * (len(term) ** 1.5)  # 長詞加權：出現次數 × 詞長^1.5
                if score > best_score:
                    best_score, best_term = score, term
        if best_term is None:  # stem fallback（只在完全無 full-match 時啟用）
            for term in vocab:
                if term and len(term) >= 2:
                    stem = term[:min(4, len(term))]
                    if stem in text:
                        best_term = term
                        break
        if best_term:
            return best_term

    # 2. Static fallback（通用主題兜底）
    static_rules = [
        ("母性保護",   ["母性", "女性勞工", "妊娠", "孕婦", "哺乳"]),
        ("過負荷預防", ["過負荷", "異常工作負荷", "促發疾病", "長工時", "疲勞"]),
        ("不法侵害",   ["不法侵害", "職場暴力", "霸凌", "騷擾"]),
        ("人因危害",   ["人因", "肌肉骨骼", "重複性", "搬運", "工效"]),
        ("消防安全",   ["消防", "排煙", "火災", "警報"]),
        ("氣體偵測",   ["氣體偵測", "氣體監控", "毒性氣體"]),
        ("洗眼沖淋",   ["洗眼", "沖淋", "急救"]),
        ("緊急應變",   ["緊急應變", "應變程序"]),
    ]
    for label, kws in static_rules:
        for kw in kws:
            if kw in text:
                return label

    # 3. Behavior 名稱作最後防線
    b = (candidate.get("behavior") or "").strip()
    return b[:30] if b else "OTHER"


def compute_topics_observed_dist(topk: list, vocab: list = None) -> dict:
    """
    topics_observed_dist（Observed）統計。
    vocab = facets_expected → 傳給 infer_topic_from_candidate 做動態比對。
    """
    dist = {}
    for c in (topk or []):
        k = infer_topic_from_candidate(c, vocab=vocab)
        k = _normalize_topic_token(k) or "OTHER"
        dist[k] = dist.get(k, 0) + 1
    return dist


def select_per_facet_candidates(candidates: list, facets: list) -> list:
    """
    多面向題：為每個 facet 選出命中分數最高的候選，回傳去重後的清單。
    採用加權字元命中（長詞優先），與 infer_topic_from_candidate 邏輯一致。
    回傳: List[candidate_dict]，每個 facet 最多 1 個，且同一 row_id 不重複。
    """
    # facet_term -> list[(candidate, score)]
    # 每個 term 收集所有命中的候選（不只最高分那一個），
    # 這樣同一題有「員關版」和「稽核室版」時兩個都能保留
    facet_candidates: dict = {}

    for c in candidates:
        matches = c.get("matches") or []
        text = " ".join([
            c.get("behavior") or "",
            c.get("article") or "",
            " ".join(m.get("keyword") or "" for m in matches),
            " ".join(m.get("impact") or "" for m in matches),
        ])
        for term in facets:
            if not term:
                continue
            cnt = text.count(term)
            if cnt > 0:
                score = cnt * (len(term) ** 1.5)
            else:
                # n-gram fallback：取 term 中長度 ≥ 3 的子字串做部分匹配
                best_ng = 0.0
                for ng_len in range(min(len(term), 5), 2, -1):
                    for start in range(len(term) - ng_len + 1):
                        ng = term[start:start + ng_len]
                        if ng in text:
                            ng_score = (ng_len ** 1.5) * 0.5  # 部分匹配打 5 折
                            if ng_score > best_ng:
                                best_ng = ng_score
                score = best_ng
            if score > 0:
                if term not in facet_candidates:
                    facet_candidates[term] = []
                cid = c.get("row_id") or str(id(c))
                existing_ids = {x[0].get("row_id") or str(id(x[0])) for x in facet_candidates[term]}
                if cid not in existing_ids:
                    facet_candidates[term].append((c, score))

    # 去重：同一 row_id 只保留一次，但同 facet 的多個候選都放入（捕捉多部門）
    # 每個 term 最多保留 top-2（避免過多雜訊），按分數排序
    seen_ids = set()
    result = []
    # 先取每個 term 的最佳分數，依此排序 term（高分的 term 優先）
    sorted_terms = sorted(
        facet_candidates.items(),
        key=lambda kv: -max(s for _, s in kv[1])
    )
    for _term, cand_list in sorted_terms:
        for c, _score in sorted(cand_list, key=lambda x: -x[1])[:2]:  # 每 facet 最多 2 個
            rid = c.get("row_id") or str(id(c))
            if rid not in seen_ids:
                seen_ids.add(rid)
                result.append(c)

    return result


def compute_facet_coverage_expected_vs_observed(facets_expected: list, topics_observed_dist: dict) -> tuple:
    """Expected facets 覆蓋到 Observed topics 的命中數。"""
    if not facets_expected:
        return 0, 0

    observed_keys = {_normalize_topic_token(k) for k in (topics_observed_dist or {}).keys()}
    hit = 0
    for facet in facets_expected:
        if _normalize_topic_token(facet) in observed_keys:
            hit += 1
    return hit, len(facets_expected)


def classify_yellow_reason_expected_vs_observed(multifacet_tag: bool,
                                               facets_expected: list,
                                               topics_observed_dist: dict) -> str:
    """黃色原因分類（Expected vs Observed）。"""
    if multifacet_tag and facets_expected:
        hit, total = compute_facet_coverage_expected_vs_observed(facets_expected, topics_observed_dist)
        if total > 0 and hit > 0:
            return "multifacet_partial_coverage"
        return "semantic_gap_or_missing_db"
    return "semantic_gap_or_missing_db"


def is_legal_clause(question: str) -> bool:
    """
    識別「合約法律條款」（違約罰則、合約效力等），DB 無對應 RBA 行為準則。

    觸發時標黃色「法律條款，請依人工比對」，不計入匹配數。
    """
    legal_keywords = [
        # 合約效力
        "签署后生效", "簽署後生效", "长期有效", "長期有效",
        "不可撤销", "不可撤銷", "取代本承诺书", "取代本承諾書",
        "业务存续期间", "業務存續期間",
        # 違約罰則
        "违约金", "違約金", "违约责任", "違約責任",
        "追究法律责任", "追究法律責任", "追究其法律",
        "径行解除", "徑行解除", "解除或终止", "解除或終止",
        "没收已发生的贿赂", "沒收已發生的賄賂",
        # 甲方權利
        "甲方有权", "甲方有權", "甲方得径行", "甲方得徑行",
        "甲方无须", "甲方無須承擔",
        # 刑事/司法
        "涉及犯罪", "依法追究", "刑事责任", "刑事責任",
        "民事责任", "民事責任",
    ]
    return any(kw in question for kw in legal_keywords)


def rerank_with_coverage_gating(candidates, must_have_terms, debug=False):
    """
    多概念覆蓋度評分（Rerank 優化）
    
    問題：多概念題目（如：非法競爭 + 腐敗 + 智財）容易只命中一個概念
    解決：檢查候選條文是否涵蓋題目的所有核心概念
    
    參數：
        candidates: RAG 返回的候選列表 [{"behavior": ..., "article": ..., "impact": ..., "score": ...}]
        must_have_terms: LLM 萃取的必備關鍵詞 ["非法競爭", "腐敗", "智慧財產權"]
        debug: 是否輸出詳細 log
    
    回傳：
        調整後的候選列表（score 已更新）
    """
    if not must_have_terms or len(must_have_terms) == 0:
        return candidates
    
    for cand in candidates:
        # 合併候選條文的所有文字內容
        full_text = " ".join([
            cand.get("behavior", ""),
            cand.get("article", ""),
            cand.get("impact", ""),
        ]).lower()
        
        # 計算命中的 must_have_terms 數量
        matched_count = sum(1 for term in must_have_terms if term.lower() in full_text)
        coverage = matched_count / len(must_have_terms) if len(must_have_terms) > 0 else 0
        
        # 覆蓋度加分/扣分
        if coverage >= 0.8:  # 涵蓋 80% 以上（如 3/3 或 5/6）
            boost = 20
        elif coverage >= 0.5:  # 涵蓋 50% 以上（如 2/3 或 3/6）
            boost = 10
        elif coverage >= 0.3:  # 涵蓋 30% 以上（如 1/3 或 2/6）
            boost = 0  # 不加分也不扣分
        else:  # 涵蓋不足 30%（如 0/3 或 1/6）
            boost = -15  # 大幅降權
        
        # 更新分數
        original_score = cand.get("best_score", cand.get("score", 0))
        cand["best_score"] = original_score + boost
        cand["coverage"] = coverage
        cand["coverage_boost"] = boost
        
        if debug:
            print(f"  [Coverage] {cand.get('behavior', '')[:30]} | matched={matched_count}/{len(must_have_terms)} | coverage={coverage:.1%} | boost={boost:+d}")
    
    # 重新排序
    candidates.sort(key=lambda x: x.get("best_score", 0), reverse=True)
    
    return candidates


def exact_token_boost(candidates, question_schema, debug=False):
    """
    精確 Token 匹配加分（ISO/標準編號/數字精準匹配）
    
    問題：ISO 14001 vs ISO 14064 embedding 相似，但實際不同
    解決：如果題目要求 ISO 14001，候選條文沒有 "14001" 就降權
    
    參數：
        candidates: RAG 返回的候選列表
        question_schema: LLM 萃取的題目 schema（QuestionItem）
        debug: 是否輸出詳細 log
    
    回傳：
        調整後的候選列表
    """
    # 從題目中提取精確 token（ISO 編號、SA 編號、數字等）
    exact_tokens = []
    
    # 1. 從 must_have_terms 中提取
    for term in question_schema.must_have_terms:
        # ISO/SA/RBA 等標準編號
        if re.search(r'(ISO|SA|RBA)\s*\d+', term, re.I):
            exact_tokens.append(term.upper())
        # 純數字編號（如 14001, 8000）
        elif re.search(r'\d{4,}', term):
            exact_tokens.extend(re.findall(r'\d{4,}', term))
    
    # 2. 從原始題目中提取（備用）
    question_text = question_schema.intent + " " + " ".join(question_schema.must_have_terms)
    iso_matches = re.findall(r'ISO\s*(\d+)', question_text, re.I)
    sa_matches = re.findall(r'SA\s*(\d+)', question_text, re.I)
    exact_tokens.extend([f"ISO{num}" for num in iso_matches])
    exact_tokens.extend([f"SA{num}" for num in sa_matches])
    
    # 去重
    exact_tokens = list(set([t.upper() for t in exact_tokens if t]))
    
    if not exact_tokens:
        return candidates  # 沒有精確 token，不做處理
    
    if debug:
        print(f"  [Exact Token] 提取到精確 token: {exact_tokens}")
    
    # 檢查每個候選條文
    for cand in candidates:
        full_text = " ".join([
            cand.get("behavior", ""),
            cand.get("article", ""),
            cand.get("impact", ""),
        ]).upper()
        
        matched_tokens = [token for token in exact_tokens if token in full_text]
        
        if len(matched_tokens) == len(exact_tokens):
            # 完全匹配所有精確 token
            boost = 30
        elif len(matched_tokens) > 0:
            # 部分匹配
            boost = 15
        else:
            # 完全沒有匹配 → 大幅降權（避免 ISO 14064 誤中 ISO 14001）
            boost = -20
        
        original_score = cand.get("best_score", cand.get("score", 0))
        cand["best_score"] = original_score + boost
        cand["exact_token_boost"] = boost
        cand["matched_tokens"] = matched_tokens
        
        if debug:
            print(f"  [Exact Token] {cand.get('behavior', '')[:30]} | matched={matched_tokens} | boost={boost:+d}")
    
    # 重新排序
    candidates.sort(key=lambda x: x.get("best_score", 0), reverse=True)
    
    return candidates


# 顏色樣式（全域共用）
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# ★ 文字樣式：D 欄用（水平+垂直置中 + 自動換行）
ALIGN_D_COLUMN = Alignment(wrap_text=True, vertical="center", horizontal="center")

# ★ 文字樣式：E 欄用（靠左+垂直置中 + 自動換行，因為內容較多）
ALIGN_E_COLUMN = Alignment(wrap_text=True, vertical="center", horizontal="left")

# ========== Azure OpenAI 配置 ==========
# ⚠️ 如果要停用 LLM，把 AZURE_OPENAI_ENABLED 改成 False 即可
AZURE_OPENAI_ENABLED = True  # 總開關（暫時關閉用於測試）

AZURE_OPENAI_ENDPOINT = "https://en-openai01.openai.azure.com"
AZURE_OPENAI_API_KEY = "1xPk95hy7EROtFp4yytQD6jrb237tGuewf0vGuMWExcbAr1TSgjUJQQJ99BFACHYHv6XJ3w3AAABACOGdXkj"
AZURE_OPENAI_DEPLOYMENT = "gpt-5.3-chat"
AZURE_OPENAI_API_VERSION = "2024-12-01-preview"

# ========== RAG 語意搜尋配置 ==========
RAG_CONFIG = {
    "base_url": "http://10.100.40.5:8004",
    "api_key": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJzdWIiOiIyM2YwNzMxNS1hN2M5LTRjMjktODQxMS02OTI2MjlkNDRmMTIiLCJpc3MiOiJuOG4iLCJhdWQiOiJwdWJsaWMtYXBpIiwiaWF0IjoxNzYwNTEyNjk5fQ.qW23MNSzD4AhAZA6yXJwvXWOSDkrHCHZS8vvG3g6jhM",  # ✅ 統一使用 RAG 服務配置的 API Key
    "pdf_name": "附件三_EnvSafety_atta3_知識庫",  # 與上傳腳本一致
    "unit": "SYSTEM",
}


def load_database(db_path):
    """
    載入附件三資料庫 Excel

    欄位對應：
      A: 類別
      B: 行為準則
      C: 關鍵字
      D: 中文條文內容
      E: 權責部門
      F: 目前現況/可能影響
      G: 問卷出處

    回傳：list[dict]
    """
    print(f"\n[資料庫載入] 開始讀取: {db_path}")

    if not os.path.exists(db_path):
        raise FileNotFoundError(f"資料庫檔案不存在: {db_path}")

    wb = openpyxl.load_workbook(db_path, data_only=True)
    ws = wb.active  # 使用第一個工作表

    print(f"[資料庫載入] 工作表名稱: {ws.title}")
    print(f"[資料庫載入] 最大行數: {ws.max_row}")

    database = []

    # 從第2行開始讀取（第1行是標題）
    for row_idx in range(2, ws.max_row + 1):
        # ✅ 資料清洗：所有欄位都先 strip() 去除前後空白
        def _clean_cell(value):
            """清洗儲存格內容：去除前後空白、None 轉空字串"""
            if value is None:
                return ""
            return str(value).strip()

        row_data = {
            "row_index": row_idx,
            "類別": _clean_cell(ws[f"A{row_idx}"].value),
            "行為準則": _clean_cell(ws[f"B{row_idx}"].value),
            "關鍵字": _clean_cell(ws[f"C{row_idx}"].value),
            "條文內容": _clean_cell(ws[f"D{row_idx}"].value),
            "權責部門": _clean_cell(ws[f"E{row_idx}"].value),
            "現況影響": _clean_cell(ws[f"F{row_idx}"].value),
            "問卷出處": _clean_cell(ws[f"G{row_idx}"].value),
        }

        # 解析關鍵字（換行分隔）
        keywords_raw = row_data["關鍵字"] or ""
        row_data["關鍵字列表"] = [
            kw.strip() for kw in keywords_raw.split("\n") if kw.strip()
        ]

        # 判斷是否為彙整版（檢查背景顏色或特殊標記）
        # 這裡暫時用簡單規則：若「條文內容」很長（>500字），視為彙整版
        content_length = len(row_data["條文內容"] or "")
        row_data["是否彙整版"] = content_length > 500

        database.append(row_data)

    wb.close()

    print(f"[資料庫載入] 成功載入 {len(database)} 筆資料")
    print(f"[資料庫載入] 彙整版資料: {sum(1 for d in database if d['是否彙整版'])} 筆")

    return database


def two_stage_matching(question, database):
    """
    兩階段比對算法

    階段 1：行為準則比對（B 欄）
      - 使用 fuzz.partial_ratio 和 fuzz.token_set_ratio
      - 閾值：60（主要）/ 50（備援）

    階段 2：關鍵字比對（C 欄）
      - 計算關鍵字重複次數
      - 支援模糊匹配

    綜合評分：
      - phase1_score * 0.4 + phase2_score * 0.6
      - 彙整版加 5 分

    回傳：best_match (dict) 或 None
    """

    if not question or not question.strip():
        return None

    # ========== 第一階段：行為準則比對 ==========
    candidates = []

    for item in database:
        behavior = item.get("行為準則") or ""
        if not behavior.strip():
            continue

        # 計算與 B 欄（行為準則）的相似度
        similarity = fuzz.partial_ratio(question, behavior)

        # 設定閾值
        threshold = MATCH_CONFIG["phase1_primary_threshold"]

        if similarity >= threshold:
            candidates.append({
                "item": item,
                "phase1_score": similarity
            })

    # 如果第一階段沒找到，降低閾值重試
    if not candidates:
        fallback_threshold = MATCH_CONFIG["phase1_fallback_threshold"]

        for item in database:
            behavior = item.get("行為準則") or ""
            if not behavior.strip():
                continue

            similarity = fuzz.token_set_ratio(question, behavior)

            if similarity >= fallback_threshold:
                candidates.append({
                    "item": item,
                    "phase1_score": similarity
                })

    # 如果還是沒找到，直接用所有資料庫項目
    if not candidates:
        candidates = [{"item": item, "phase1_score": 0} for item in database]

    # ========== 第二階段：關鍵字比對 ==========
    best_match = None
    max_score = 0

    # 對問題進行分詞
    question_words = set(jieba.cut(question))

    for candidate in candidates:
        item = candidate["item"]
        keywords = item["關鍵字列表"]

        # 計算關鍵字重複次數
        keyword_count = 0
        matched_keywords = []

        for keyword in keywords:
            if not keyword:
                continue

            # 檢查關鍵字是否出現在問題中
            if keyword in question:
                keyword_count += 1
                matched_keywords.append(keyword)
            # 使用模糊比對處理變體
            elif any(fuzz.ratio(keyword, word) > 85 for word in question_words):
                keyword_count += 0.8  # 模糊匹配給予較低分數
                matched_keywords.append(f"{keyword}(模糊)")

        # ========== 綜合評分 ==========
        phase1_score = candidate["phase1_score"]
        phase2_score = keyword_count * 10  # 每個關鍵字 10 分

        behavior_weight = MATCH_CONFIG["behavior_weight"]
        keyword_weight = MATCH_CONFIG["keyword_weight"]

        total_score = phase1_score * behavior_weight + phase2_score * keyword_weight

        # 彙整版加分（如果分數相同，優先選彙整版）
        if item["是否彙整版"]:
            total_score += MATCH_CONFIG["aggregate_bonus"]

        # 更新最佳匹配
        if total_score > max_score:
            max_score = total_score
            best_match = {
                **item,
                "confidence": total_score,
                "matched_keywords": matched_keywords,
                "phase1_score": phase1_score,
                "phase2_score": phase2_score,
                "keyword_count": keyword_count
            }

    return best_match


def azure_llm_upgrade_yellow(question, candidate):
    """
    使用 Azure OpenAI 審查黃色區間（55-70）的配對，決定是否升級成綠色。

    【角色定位】LLM 是「加分助教」，不是「裁判」：
      - 只能把黃色升級成綠色
      - 不能把綠色/黃色降級成紅色或未匹配

    參數：
      question: 客戶問卷題目
      candidate: dict，包含 behavior, category, article 等欄位（RAG top-1）

    回傳：
      True  → LLM 認為配對合理，可以升級成綠色
      False → LLM 認為配對可疑，保持黃色
      None  → 沒有啟用 / 呼叫失敗 / 回覆無法解析，保持黃色
    """
    # ★ print-L0：檢查總開關
    if not AZURE_OPENAI_ENABLED:
        print("[LLM] Azure OpenAI 已停用，略過 LLM 升級審查")
        return None

    if not (AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY and AZURE_OPENAI_DEPLOYMENT):
        print("[LLM] Azure OpenAI 配置不完整，略過 LLM 升級審查")
        return None

    try:
        behavior = candidate.get("behavior") or "無"
        category = candidate.get("category") or "無"
        article = (candidate.get("article") or "無")[:400]  # 節錄避免 token 過多
        dept = candidate.get("dept") or "無"

        system_prompt = (
            "你是一位負責 RBA 問卷題目比對的專家。"
            "你的任務是審查「語意相似但信心分數在黃色區間（55-70）」的配對，"
            "判斷它是否足夠合理，可以升級成綠色（高信心）。"
            "請根據語意契合度、主題一致性來判斷，不要過於嚴苛。"
        )

        user_content = (
            f"客戶問卷題目：\n{question}\n\n"
            f"RAG 系統找到的配對（相似度在 55-70% 黃色區間）：\n"
            f"  類別: {category}\n"
            f"  行為準則: {behavior}\n"
            f"  權責部門: {dept}\n"
            f"  條文內容（節錄）: {article}\n\n"
            "這個配對是否足夠合理，可以升級成綠色（高信心）？\n"
            "- 如果你認為配對大致合理、語意相符，請回覆 'Y'\n"
            "- 如果你認為配對有疑慮、不夠精確，請回覆 'N'\n\n"
            "只需回覆 'Y' 或 'N'，不要有其他說明。"
        )

        # 使用與 server_chatbot 相同的 Azure OpenAI 調用方式
        url = (
            f"{AZURE_OPENAI_ENDPOINT}/openai/deployments/"
            f"{AZURE_OPENAI_DEPLOYMENT}/chat/completions"
        )

        headers = {
            "api-key": AZURE_OPENAI_API_KEY,  # Azure 用 api-key 不是 Authorization
            "Content-Type": "application/json",
        }

        payload = {
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_content},
            ],
            "seed": 42,                     # P0 Fix: 降低隨機抖動
            "max_completion_tokens": 200,   # P0 Fix: Y/N 回覆只需 200 tokens
        }

        # ★ print-L1：在呼叫前印出
        print(f"[LLM] 呼叫 Azure OpenAI 審查黃色配對是否可升級...")

        resp = requests.post(
            url,
            headers=headers,
            json=payload,
            params={"api-version": AZURE_OPENAI_API_VERSION},
            timeout=30
        )
        resp.raise_for_status()
        data = resp.json()

        content = (
            data.get("choices", [{}])[0]
            .get("message", {})
            .get("content", "")
            .strip()
            .upper()
        )

        # ★ print-L2：印出 LLM 回覆
        print(f"[LLM] 回覆內容: {content}")

        # 解析回覆
        if content.startswith("Y"):
            print(f"[LLM] ✅ LLM 認為配對合理，升級成綠色")
            return True
        elif content.startswith("N"):
            print(f"[LLM] ⚠️  LLM 認為配對有疑慮，保持黃色")
            return False
        else:
            print(f"[LLM] ⚠️  無法解析 LLM 回覆，保持黃色")
            return None

    except Exception as e:
        # ★ print-L3：印出錯誤
        print(f"[LLM] 呼叫 Azure 失敗，略過 LLM 升級審查: {repr(e)}")
        return None


def extract_field(text: str, field_name: str) -> str:
    """
    從結構化文字中提取欄位內容（支援單行和多行 RAG chunk 格式）。

    格式一（單行，RAG API 實際回傳）：
      [類別] xxx [行為準則] yyy [關鍵字] zzz [條文內容] aaa [權責部門] bbb [目前現況] ccc

    格式二（多行）：
      [行為準則] ...
      [關鍵字] ...
      [權責部門] ...
      [目前現況] 第一行
      第二行
      [問卷出處] ...

    修正：改用 field_name 精確定位，不再用第一個 ']' 截斷，
    避免單行格式時所有欄位都截到相同錯誤值。
    """
    if field_name not in text:
        return ""

    lines = text.split("\n")
    for i, line in enumerate(lines):
        if field_name not in line:
            continue

        # 從 field_name 結束後開始取值（精確定位，不用 split(']')）
        idx = line.find(field_name)
        after = line[idx + len(field_name):]

        content_lines = []

        # 單行格式：到下一個 " [" (空格+左括號，代表新欄位) 為止
        m = re.search(r'\s\[', after)
        if m:
            first_line = after[:m.start()].strip()
        else:
            first_line = after.strip()

        if first_line:
            content_lines.append(first_line)

        # 多行格式：繼續讀後續行，直到遇到新 [TAG] 行
        for j in range(i + 1, len(lines)):
            next_line = lines[j].strip()
            if next_line.startswith("["):
                break
            if next_line:
                content_lines.append(next_line)

        return "\n".join(content_lines)

    return ""

def parse_no_skip_target(question: str) -> str:
    """
    從題目文字解析「NO 分支要跳到哪一題」
    支援中英文：
      - If no, please skip to C.7.1
      - 若否，請直接跳至 C.7.1
    回傳例如 "C.7.1"，找不到回傳 ""。
    """
    import re
    q = (question or "").strip()
    if not q:
        return ""

    # 先抓英文 if no ... skip to ...
    m = re.search(
        r"if\s*no[\s\S]{0,80}?(?:skip\s*to|go\s*to)\s*(?:the\s*)?([a-z]\.\d+(?:\.\d+)*)",
        q,
        flags=re.IGNORECASE
    )
    if m:
        return m.group(1).upper()

    # 再抓中文 若否 ... 跳至/跳到 ...
    m = re.search(
        r"(?:若否|否則)[\s\S]{0,40}?(?:跳至|跳到|直接跳至)\s*([a-z]\.\d+(?:\.\d+)*)",
        q,
        flags=re.IGNORECASE
    )
    if m:
        return m.group(1).upper()

    return ""


def is_answer_no(impact_text: str) -> bool:
    """
    判斷答案是否為「NO」或否定含義
    檢查現況/影響欄位是否包含：
      - NO (英文)
      - 否/無 (中文)
      - 未執行/無執行/未設定/無設定
    """
    if not impact_text:
        return False

    text_lower = impact_text.lower().strip()

    # 英文否定詞（獨立單字或開頭）
    if text_lower.startswith("no"):
        return True
    if ",no," in text_lower or " no," in text_lower or " no " in text_lower:
        return True

    # 中文否定詞
    negative_patterns = [
        "否", "無", "未執行", "無執行", "未設定", "無設定",
        "不執行", "不設定", "沒有", "未有", "無此", "未建立"
    ]

    for pattern in negative_patterns:
        if pattern in impact_text:
            return True

    return False


def verify_keyword_match(question: str, candidate_text: str, min_match: int = 1) -> bool:
    """
    最小版關鍵字驗證（不依賴 jieba）：
    - 英文：抓 a-z 單字（>=3）
    - 中文：抓連續中文字詞（>=2）
    - 去掉停用詞後比交集
    """
    import re
    q = (question or "").lower()
    t = (candidate_text or "").lower()

    stop_words = {
        # 中文常見虛詞/問卷語助詞
        "的", "是", "否", "若", "請", "公司", "提供", "說明", "回覆", "直接", "是否", "並", "及", "或",
        "包含", "若是", "是否", "及報", "報表", "管理", "內容", "政策", "程序", "進行",
        # 英文常見虛詞/問卷模板字
        "does", "do", "the", "company", "if", "yes", "no", "please", "provide",
        "complete", "skip", "to", "and", "or", "once", "where", "once", "required",
        "have", "has", "with", "for", "from", "that", "this", "which", "such",
    }

    def _tokens(s: str) -> set:
        # 英文 token（>=3 個字母）
        en = set(w for w in re.findall(r"[a-z]{3,}", s) if w not in stop_words)

        # ✅ 中文 token（用滑動窗口提取 2-8 字詞，支援「緊急應變演練」等長詞）
        zh = set()
        chinese_chars = re.findall(r"[\u4e00-\u9fff]+", s)
        for seg in chinese_chars:
            # ✅ 修改：從 2-4 字擴展到 2-8 字（覆蓋率從 51.9% → 88.9%）
            for length in [2, 3, 4, 5, 6, 7, 8]:
                for i in range(len(seg) - length + 1):
                    word = seg[i:i+length]
                    if word not in stop_words:
                        zh.add(word)

        return en | zh

    qk = _tokens(q)
    tk = _tokens(t)
    common = qk & tk

    debug = os.getenv("RAG_SKIP_DEBUG", "0") == "1"
    if debug:
        # [print-KV1] ✅ 加在「下面」：關鍵字驗證細節
        print(f"[print-KV1][下面] q_keywords={sorted(list(qk))[:20]} ...")
        print(f"[print-KV1][下面] t_keywords={sorted(list(tk))[:20]} ...")
        print(f"[print-KV1][下面] common={sorted(list(common))} (n={len(common)})")

    # [KW-VERIFY-DEBUG] 弱匹配偵測：common tokens 的最長長度 < 3 = 僅靠短詞通過
    passed = len(common) >= min_match
    if passed and common:
        max_tok_len = max(len(t) for t in common)
        if max_tok_len <= 2:
            print(f"  [KW-VERIFY-WEAK] 弱匹配警告：common={sorted(list(common))} max_len={max_tok_len} (僅2字短詞命中)")
    return passed


def expand_question_for_rag(question: str) -> str:
    """
    Query Expansion：偵測子概念（fire/first aid/...）時補上同義詞/相關詞
    目的：提高 RAG 召回率（讓 row_id=94 消防、106 急救有機會進 top-k）
    """
    q_lower = (question or "").lower()

    # 子概念 → 補詞表（中英文同義詞）
    expansion_rules = [
        {
            "triggers": ["fire", "alarm", "extinguisher", "detection", "firefight", "火警", "警報", "滅火", "探測", "消防"],
            "add_terms": "火警探測器 警報系統 滅火器 消防設備 fire detection alarm system firefighting equipment extinguisher"
        },
        {
            "triggers": ["first aid", "aid kit", "medical kit", "急救", "急救箱", "醫療箱", "救護"],
            "add_terms": "急救箱 急救設備 醫療用品 first aid kit first aid equipment medical supplies"
        },
        {
            # ⚠️ 危險撤離相關 - 避免「water withdrawal」誤觸發
            "triggers": ["remove themselves", "imminent harm", "imminent danger", "danger withdraw", "退避", "危險遠離", "安全場所"],
            "excludes": ["water", "energy", "emission", "waste", "用水", "能源", "排放", "廢棄"],  # 排除環境類
            "add_terms": "退避權 遠離危險 安全場所 撤離 withdraw from danger safe place evacuation right"
        },
        {
            "triggers": ["drill", "exercise", "evacuation drill", "emergency drill", "演練", "演習", "疏散演練", "緊急演練"],
            "add_terms": "緊急疏散演練 應急演習 消防演練 逃生演練 emergency evacuation drill emergency preparedness drill fire drill"
        },
        {
            "triggers": ["risk identification", "risk assessment", "health safety risk", "hazard identification", "危害辨識", "風險識別", "風險評估", "健康安全風險"],
            "add_terms": "危害辨識 風險識別 風險鑑別 風險評估 潛在危害風險識別 健康安全風險評估 職業健康安全風險 risk identification hazard identification risk assessment occupational health safety risk"
        },
        {
            "triggers": ["identification", "assessment", "identify", "evaluate"],
            "add_terms": "識別 辨識 鑑別 評估 辨別 assessment identification evaluation"
        },
        {
            # ✅ 用水/水資源管理相關
            "triggers": ["water withdrawal", "water usage", "water reduction", "water resource", "用水減量", "取水", "水資源", "節水"],
            "add_terms": "用水管理 節水製程 水資源管理計畫 用水減量目標 節水機會 water management water conservation water resource management"
        },
    ]

    expanded = question
    for rule in expansion_rules:
        # 檢查觸發條件
        if any(t in q_lower for t in rule["triggers"]):
            # ✅ 檢查排除條件（避免誤觸發）
            excludes = rule.get("excludes", [])
            if excludes and any(ex in q_lower for ex in excludes):
                continue  # 跳過此規則

            # 只補一次（避免重複累積）
            if rule["add_terms"] not in expanded:
                expanded = expanded + " " + rule["add_terms"]

    return expanded


def rerank_rag_candidates(question: str, candidates: list, top_n: int = 8) -> dict:
    """
    在 RAG top-k 候選內做「先C再D + 衝突扣分」二次排序（不依賴 jieba）
    + 新增 must-have gate：題目有子概念時，候選必須命中核心詞，否則拒答（回 None）
    """
    import re
    import os

    if not candidates:
        return None

    # ===== 權重（依你的指定）=====
    c_hit_weight = 12
    d_hit_weight = 5
    conflict_penalty = 35

    # ✅ must-have 缺失扣分（足夠大，讓泛用條文出局）
    must_have_penalty = 120

    q_raw = (question or "").strip()
    q = q_raw.lower()
    cand_list = candidates[:max(1, top_n)]

    stop = {
        # 中文常見虛詞/問卷語助詞
        "的", "是", "否", "若", "請", "公司", "提供", "說明", "回覆", "直接", "是否", "並", "及", "或",
        # 英文常見虛詞/問卷模板字
        "does", "do", "the", "company", "if", "yes", "no", "please", "provide",
        "complete", "skip", "to", "and", "or", "once", "where", "required",
        "have", "has", "with", "for", "from", "that", "this", "which", "such",
    }

    # 模板詞（越多越像「泛用制度」）→ 小扣分
    generic_zh = {"制定", "流程", "管理", "監控", "確保", "政策", "程序", "機制", "辦法", "規範", "作業"}
    generic_en = {"procedure", "process", "management", "ensure", "policy", "monitor", "mechanism", "establish"}

    # ✅ 子概念規則表（重新設計：每類題目有專屬規則，避免互相干擾）
    SUBCONCEPT_RULES = [
        {
            "name": "fire_equipment",  # 消防設備（B.2.1）
            "triggers": ["fire detection", "fire alarm", "fire extinguisher", "firefighting equipment",
                        "火災偵測", "火警探測", "滅火器", "滅火設備", "消防設備"],
            "must_have_any": ["fire", "alarm", "extinguisher", "detection", "firefighting",
                             "火警", "警報", "滅火", "探測", "消防"],
            "conflicts": ["drill", "exercise", "演練", "演習"],  # 排除演練類
        },
        {
            "name": "emergency_drill",  # 緊急演練（B.2.3）
            "triggers": ["evacuation drill", "emergency drill", "fire drill",
                        "疏散演練", "緊急演練", "消防演練", "逃生演練"],
            "must_have_any": ["drill", "exercise", "演練", "演習"],
            "conflicts": [],  # 演練類題目不與其他概念衝突
        },
        {
            "name": "first_aid_equipment",  # 急救設備（B.3.3）
            "triggers": ["first aid kit", "first aid equipment", "medical equipment",
                        "急救箱", "急救設備", "醫療設備"],
            "must_have_any": ["first aid", "aid kit", "medical equipment", "急救", "急救箱"],
            "conflicts": ["drill", "exercise", "evacuation", "演練", "演習", "疏散"],  # 排除演練類
        },
        {
            "name": "withdraw_danger",  # 遠離危險（B.3.1）
            # ⚠️ 避免「water withdrawal」（取水）誤觸發，只用精確短語
            "triggers": ["remove themselves", "imminent danger", "imminent harm",
                        "safe place", "danger withdraw",
                        "危險", "遠離危險", "退避", "安全場所"],
            "must_have_any": ["danger", "safe", "remove", "imminent", "harm",
                             "危險", "退避", "安全", "遠離危險"],
            "conflicts": ["grievance", "complaint", "anonymous", "water", "energy",
                         "申訴", "檢舉", "匿名", "取水", "用水", "能源"],  # 排除申訴類和環境類
        },
    ]

    def toks(s: str) -> dict:
        """
        ✅ 修改：返回 dict {token: weight} 而非 set
        長詞權重更高，避免「應變演練」(4字) 贏過「緊急應變演練」(6字)
        """
        s = (s or "").lower()
        # 英文 token（權重固定為 3）
        en = {w: 3 for w in re.findall(r"[a-z]{3,}", s) if w not in stop}

        # ✅ 中文 token（2-8 字滑窗，長詞權重更高）
        zh = {}
        for seg in re.findall(r"[\u4e00-\u9fff]+", s):
            # ✅ 修改：從 2-4 字擴展到 2-8 字
            for L in (2, 3, 4, 5, 6, 7, 8):
                for i in range(len(seg) - L + 1):
                    w = seg[i:i + L]
                    if w not in stop:
                        # ✅ 長詞加權：權重 = 長度
                        # 例如：「緊急應變演練」(6字)權重6，「應變演練」(4字)權重4
                        if w not in zh or zh[w] < L:
                            zh[w] = L

        return {**en, **zh}

    def count_generic(s: str) -> int:
        s2 = (s or "").lower()
        c = 0
        for w in generic_zh:
            if w in s2:
                c += 1
        for w in generic_en:
            if w in s2:
                c += 1
        return c

    # ===== 偵測本題屬於哪些子概念（可能多個）=====
    active_rules = []
    for r in SUBCONCEPT_RULES:
        if any(t.lower() in q for t in r["triggers"]):
            active_rules.append(r)

    qk = toks(q)

    def conflict_hits(candidate_text: str) -> int:
        ct = (candidate_text or "").lower()
        hits = 0
        for r in active_rules:
            if any(x in ct for x in r.get("conflicts", [])):
                hits += 1
        return hits

    def must_have_missing(candidate_text: str) -> int:
        """
        若本題有子概念，候選必須命中 must_have_any 的任一詞；否則視為缺失
        """
        if not active_rules:
            return 0
        ct = (candidate_text or "").lower()
        missing = 0
        for r in active_rules:
            must_any = r.get("must_have_any", [])
            if must_any and not any(x.lower() in ct for x in must_any):
                missing += 1
        return missing

    best = None
    best_score = -1e9

    # 用於判斷：是否所有候選都缺 must-have（缺到應該整題未匹配）
    all_missing = True

    for c in cand_list:
        base = float(c.get("best_score", 0.0))  # 0~1
        category = c.get("category") or ""
        behavior = c.get("behavior") or ""
        article = c.get("article") or ""
        matches = c.get("matches") or []

        kw_text = " ".join([(m.get("keyword") or "") for m in matches])
        c_text = " ".join([category, behavior, kw_text])

        ck = toks(c_text)
        dk = toks(article)

        # ✅ 加權計算：優先選擇長詞
        # 例如：若 qk 和 ck 都有「緊急應變演練」(權重6)和「應變演練」(權重4)
        #      則取權重較高的「緊急應變演練」
        c_hit_score = sum(max(qk.get(w, 0), ck.get(w, 0)) for w in (set(qk.keys()) & set(ck.keys())))
        d_hit_score = sum(max(qk.get(w, 0), dk.get(w, 0)) for w in (set(qk.keys()) & set(dk.keys())))

        # 轉換為等效的「命中數」（除以平均權重 4）
        c_hit = c_hit_score / 4.0
        d_hit = d_hit_score / 4.0

        # ✅ 壓縮 + 上限：避免 D 欄長文本把分數炸穿
        c_hit_eff = min(c_hit, 12)                 # C 欄最多吃到 12
        d_hit_eff = min(d_hit, 20)                 # D 欄最多吃到 20（避免 170 這種情況）
        c_bonus = (c_hit_eff ** 0.5) * c_hit_weight  # sqrt 壓縮
        d_bonus = (d_hit_eff ** 0.5) * d_hit_weight  # sqrt 壓縮

        # 泛用模板詞扣分（輕微）
        penalty_generic = count_generic(" ".join([behavior, article]))
        generic_penalty_score = penalty_generic * 2.5

        # 子項衝突扣分（強）
        full_text = " ".join([behavior, article, kw_text])
        conflict_n = conflict_hits(full_text)
        conflict_penalty_score = conflict_n * conflict_penalty

        # ✅ must-have gate（核心概念缺失 → 重扣）
        missing_n = must_have_missing(full_text)
        missing_penalty_score = missing_n * must_have_penalty
        if missing_n == 0:
            all_missing = False

        # ✅ 核心關鍵字精確匹配加分（避免長文本淹沒精準答案）
        keyword_match_bonus = 0
        core_keywords = {
            # 題目核心詞 → 知識庫關鍵字
            "風險": ["風險識別", "危害辨識", "風險評估", "健康安全風險評估"],
            "識別": ["風險識別", "危害辨識", "風險評估", "危害風險識別"],
            "assessment": ["風險識別", "危害辨識", "風險評估", "健康安全風險評估"],
            "identification": ["風險識別", "危害辨識", "風險評估"],
            "消防": ["消防安全設備", "滅火系統", "火警探測"],
            "fire": ["消防安全設備", "滅火系統", "火警探測"],
            "急救": ["急救設備", "急救箱"],
            "first aid": ["急救設備", "急救箱"],
            # ✅ 新增：6-8 字長詞（避免被 4 字詞搶走分數）
            "緊急應變演練": ["緊急應變演練", "緊急疏散演練", "應變演練"],
            "緊急疏散演練": ["緊急疏散演練", "緊急應變演練"],
            "emergency drill": ["緊急應變演練", "緊急疏散演練"],
            "evacuation drill": ["緊急疏散演練"],
            "資安演練": ["資安演練", "資訊安全演練"],
            "演練": ["緊急疏散演練", "消防演練", "逃生演練", "緊急應變演練"],
            "drill": ["緊急疏散演練", "消防演練", "逃生演練", "緊急應變演練"],
            # ✅ 新增：環境緊急應變
            "環境緊急應變": ["環境緊急應變計畫", "環境緊急應變"],
            "emergency response": ["緊急應變", "環境緊急應變計畫"],
            # ✅ 新增：職業安全衛生
            "職業安全衛生": ["職業安全衛生訓練", "職業安全衛生風險評估"],
            "撤離": ["人權", "退避權", "危險遠離"],
            "withdraw": ["人權", "退避權", "危險遠離"],
            # ✅ 人權保護特徵 - B.3.1 專用
            "報復": ["人權"],
            "恐嚇": ["人權"],
            "懲罰": ["人權"],
            "處罰": ["人權"],
            "retaliation": ["人權"],
            "fear": ["人權"],
        }

        # 檢查題目是否包含核心詞，候選項關鍵字是否精確匹配
        for q_keyword, kb_keywords in core_keywords.items():
            if q_keyword.lower() in q:
                for kb_kw in kb_keywords:
                    if kb_kw in kw_text:
                        keyword_match_bonus += 15  # 每個核心關鍵字匹配加 15 分
                        break  # 同一個 q_keyword 只加一次分

        score = (
            base * 100.0
            + c_bonus
            + d_bonus
            + keyword_match_bonus
            - generic_penalty_score
            - conflict_penalty_score
            - missing_penalty_score
        )

        if score > best_score:
            best_score = score
            best = c

        if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
            # [print-RR0] ✅ 加在「下面」：逐候選評分
            row_id = (c.get("row_id") or "")
            print(
                f"[print-RR0][下面] row_id={row_id} base={base*100:.1f} c_hit={c_hit} d_hit={d_hit} "
                f"keyword_bonus={keyword_match_bonus:.1f} "
                f"generic_pen={generic_penalty_score:.1f} conflict_n={conflict_n} conflict_pen={conflict_penalty_score:.1f} "
                f"missing_n={missing_n} missing_pen={missing_penalty_score:.1f} score={score:.1f} "
                f"behavior={(behavior[:28] + '...') if len(behavior) > 28 else behavior}"
            )
            # [print-RR3] ✅ 加在「下面」：用來確認 cap/壓縮是否生效
            print(f"[print-RR3][下面] c_hit={c_hit}→{c_hit_eff} d_hit={d_hit}→{d_hit_eff} c_bonus={c_bonus:.1f} d_bonus={d_bonus:.1f}")

    # ✅ 若本題屬於子概念（fire/first aid...）但所有候選都缺核心詞：直接拒答（回 None）
    if active_rules and all_missing:
        if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
            # [print-RR1] ✅ 加在「下面」
            names = ",".join([r["name"] for r in active_rules])
            print(f"[print-RR1][下面] rerank_reject_all_missing rules={names} -> return None")
        return None

    if os.getenv("RAG_SKIP_DEBUG", "0") == "1" and best:
        # [print-RR1] ✅ 加在「下面」：rerank 最終 winner
        print(
            f"[print-RR1][下面] rerank_winner best_score={best_score:.1f} "
            f"best_base={float(best.get('best_score', 0))*100:.1f} "
            f"best_behavior={((best.get('behavior') or '')[:40])}..."
        )

    return best


def get_latest_pdf_id():
    """
    動態查詢資料庫，取得最新的知識庫 PDF ID

    策略：
      1. 優先查詢資料庫（動態查詢最新的知識庫）
      2. 如果查詢失敗，fallback 到讀取 rag_config.txt
      3. 如果都失敗，返回 None（RAG 會使用 unit 過濾）

    回傳：pdf_id (str) 或 None
    """
    from sqlmodel import create_engine
    from sqlalchemy import text

    DB_URL = "postgresql://dgtk:dgtk@10.100.40.5:8002/dgtk"

    # 策略 1: 動態查詢資料庫
    try:
        print(f"  [PDF_ID] 動態查詢資料庫中的最新知識庫...")
        engine = create_engine(DB_URL, echo=False)

        with engine.connect() as conn:
            query = text("""
                SELECT id FROM pdffile
                WHERE name = :pdf_name
                AND unit = :unit
                ORDER BY date DESC
                LIMIT 1
            """)

            result = conn.execute(query, {
                "pdf_name": RAG_CONFIG["pdf_name"],
                "unit": RAG_CONFIG["unit"]
            })
            row = result.fetchone()

            if row:
                pdf_id = row[0]
                print(f"  [PDF_ID] ✅ 找到最新知識庫: {pdf_id}")
                return pdf_id
            else:
                print(f"  [PDF_ID] ⚠️  資料庫中未找到知識庫 '{RAG_CONFIG['pdf_name']}'")

    except Exception as e:
        print(f"  [PDF_ID] ⚠️  資料庫查詢失敗: {repr(e)}")

    # 策略 2: fallback 到讀取配置檔案
    try:
        config_path = "/home/lladm/frank/n8n-MCP/客戶問卷RPA/rag_config.txt"
        print(f"  [PDF_ID] 嘗試讀取配置檔案: {config_path}")

        with open(config_path, "r", encoding="utf-8") as f:
            for line in f:
                if line.startswith("PDF_ID="):
                    pdf_id = line.split("=", 1)[1].strip()
                    print(f"  [PDF_ID] ✅ 從配置檔案讀取: {pdf_id}")
                    return pdf_id

    except Exception as e:
        print(f"  [PDF_ID] ⚠️  配置檔案讀取失敗: {repr(e)}")

    # 策略 3: 全部失敗，拋出例外（避免用錯知識庫產生錯誤結果）
    # 策略 3: 全部失敗，拋出例外（避免用錯知識庫產生錯誤結果）
    pdf_name_cfg = RAG_CONFIG["pdf_name"]
    unit_cfg = RAG_CONFIG["unit"]
    raise RuntimeError(
        f"無法取得附件三知識庫 PDF ID（name={pdf_name_cfg!r}, unit={unit_cfg!r}）。"
        " 請先至 /esg/upload/ 上傳附件三知識庫後再執行 RPA。"
    )
SUBCONCEPT_KEYWORDS = {
    "fire_safety": [
        "fire", "alarm", "detection", "extinguisher", "firefighting", "evacuation",
        "火警", "警報", "探測", "滅火", "滅火器", "消防", "疏散", "逃生"
    ],
    "first_aid": [
        "first aid", "aid kit", "medical", "medical supplies",
        "急救", "急救箱", "醫療", "救護", "醫療用品"
    ],
    "withdraw_danger": [
        "withdraw", "danger", "safe place", "imminent harm", "retaliation", "return",
        "退避", "危險", "安全場所", "撤離", "報復", "返回", "遠離"
    ],
}


def _detect_active_concepts(expanded_query: str) -> list:
    """
    偵測題目屬於哪些子概念（fire_safety / first_aid / withdraw_danger）
    回傳：list of concept names
    """
    q = (expanded_query or "").lower()
    active = []

    # 火災安全
    if any(t in q for t in ["fire", "alarm", "detection", "extinguisher", "firefight"]) or \
       any(t in expanded_query for t in ["火警", "警報", "探測", "滅火", "消防"]):
        active.append("fire_safety")

    # 急救
    if any(t in q for t in ["first aid", "aid kit", "medical kit"]) or \
       any(t in expanded_query for t in ["急救", "急救箱", "醫療", "救護"]):
        active.append("first_aid")

    # 退避危險
    if any(t in q for t in ["withdraw", "danger", "safe place", "imminent harm"]) or \
       any(t in expanded_query for t in ["退避", "危險", "安全場所", "撤離"]):
        active.append("withdraw_danger")

    return active


def _norm_text(s) -> str:
    """正規化文字：去除多餘空白、統一格式"""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


# ========== Row ID 提取與正規化（修正 embedding 候選 row_id 空白問題）==========

ROW_ID_RE = re.compile(r"\brow_id\s*[:=]\s*([A-Za-z0-9_-]+)\b", re.IGNORECASE)
ROW_ID_BRACKET_RE = re.compile(r"\[row_id\]\s*([A-Za-z0-9_-]+)\b", re.IGNORECASE)

def _extract_row_id_from_text(text: str) -> str:
    """從文本中提取 row_id（多種模式匹配）"""
    if not text:
        return ""
    # 優先匹配 [row_id] A69 格式
    m = ROW_ID_BRACKET_RE.search(text)
    if m:
        return (m.group(1) or "").strip()
    # 備用匹配 row_id:A69 或 row_id=A69 格式
    m = ROW_ID_RE.search(text)
    if m:
        return (m.group(1) or "").strip()
    return ""


def normalize_rag_item(item: dict, fallback_prefix: str = "EMB") -> dict:
    """
    將 RAG/keyword 的原始 item 正規化成統一 candidate 格式。
    修正點：row_id 多路 fallback，避免 embedding 候選 row_id 空白導致不可追蹤。

    參數：
        item: RAG API 或 keyword search 返回的原始字典
        fallback_prefix: 當完全無法提取 row_id 時的前綴（EMB=embedding, KW=keyword）

    回傳：正規化後的 candidate 字典
    """
    text = (item.get("text") or item.get("content") or item.get("clause") or "").strip()

    # ① 先嘗試直接欄位
    row_id = (item.get("row_id") or "").strip()

    # ② 常見替代欄位（看你 RAG API 回什麼：chunk_id / id / parent_id）
    if not row_id:
        row_id = (item.get("chunk_id") or item.get("id") or item.get("parent_id") or "").strip()

    # ③ 從 text 內解析（你 ingest 若有把 row_id 寫進內容，這裡就能抓回來）
    if not row_id:
        row_id = _extract_row_id_from_text(text)

    score = float(item.get("score", 0.0) or 0.0)
    behavior = (item.get("behavior") or item.get("rule") or item.get("category") or "").strip()

    cand = {
        "row_id": row_id,           # 這裡不再容易是空
        "score": score,
        "behavior": behavior,
        "text": text,
        "source": (item.get("source") or "").strip(),  # 可選：保留來源
        "raw": item,                                     # 可選：保留原始方便 debug
    }

    # ④ 最後保險：仍空就給一個 trace id（至少 debug 可追）
    if not cand["row_id"]:
        # 注意：不要覆蓋真正 row_id，只在完全抓不到時才給
        cand["row_id"] = f"{fallback_prefix}_UNKNOWN"

    return cand


def build_keyword_search_terms(active_concepts: list, must_keywords: list) -> list:
    """
    keyword 補撈用詞收斂：
    - must_keywords：給 must-have gate 用（判斷缺詞/扣分/拒答）
    - search_terms：給 keyword retrieval 用（避免太泛用的詞把候選洗歪）

    參數：
        active_concepts: 偵測到的子概念列表（如 ["withdraw_danger"]）
        must_keywords: 原始必備關鍵字列表

    回傳：收斂後的搜尋詞列表
    """
    terms = list(must_keywords or [])

    # withdraw_danger 特別處理：移除過泛用的「報復/返回」相關詞
    if "withdraw_danger" in active_concepts:
        drop = {"retaliation", "return", "報復", "返回"}
        terms = [t for t in terms if (t or "").strip() and (t.lower() not in drop) and (t not in drop)]

        # 強化錨點（確保真的拉得到退避/撤離文本）
        anchor = ["withdraw", "danger", "safe place", "imminent harm", "退避", "危險", "安全場所", "撤離"]
        for a in anchor:
            if a not in terms:
                terms.append(a)

    return terms


def _keyword_search_attachment3(database: list, keywords: list, limit: int = 10) -> list:
    """
    從附件三 database (list[dict]) 做 keyword 補撈

    參數：
        database: load_database() 回傳的 list[dict]
        keywords: 要搜尋的關鍵字列表
        limit: 最多回傳幾筆

    回傳：list[dict] 格式與 RAG 返回格式相容
    """
    if not database or not keywords:
        return []

    # 過濾有效關鍵字（至少 2 個字元）
    kws = [k for k in keywords if k and len(k.strip()) >= 2]
    if not kws:
        return []

    rows = []

    for row in database:
        # 合併所有可搜尋欄位
        searchable_text = " | ".join([
            _norm_text(row.get("類別")),
            _norm_text(row.get("行為準則")),
            _norm_text(row.get("關鍵字")),
            _norm_text(row.get("條文內容")),
            _norm_text(row.get("權責部門")),
        ])

        searchable_lower = searchable_text.lower()

        # 計算關鍵字命中數
        hit_count = 0
        for kw in kws:
            kw_lower = kw.lower()
            if kw_lower in searchable_lower:
                hit_count += 1

        # 有命中才加入候選
        if hit_count > 0:
            # 給予 pseudo score（基於命中數，範圍 70-96）
            # [方案二-A] pseudo_score 校正到與 semantic 同量級 (0.08~0.14)
            # 原公式 0.70~0.96 會壓過語意分數(0.07~0.15)，造成排序失真
            pseudo_score = min(0.14, 0.08 + hit_count * 0.01)

            # 轉換成與 RAG 回傳格式相容的結構
            rows.append({
                "row_id": f"A{row.get('row_index', '?')}",
                "article": _norm_text(row.get("條文內容")),
                "category": _norm_text(row.get("類別")),
                "behavior": _norm_text(row.get("行為準則")),
                "best_score": pseudo_score,
                "matches": [{
                    "row_id": f"A{row.get('row_index', '?')}",
                    "keyword": _norm_text(row.get("關鍵字")),
                    "dept": _norm_text(row.get("權責部門")),
                    "impact": _norm_text(row.get("現況影響")),
                    "score": pseudo_score,
                }],
                "_source": "keyword_local",
                "_hit_count": hit_count,
            })

    # 按命中數和分數排序
    rows.sort(key=lambda x: (x["_hit_count"], x["best_score"]), reverse=True)

    return rows[:limit]


def _merge_dedup_candidates(embedding_results: list, keyword_results: list) -> list:
    """
    合併去重：以 row_id 當 key

    策略：
      1. embedding 結果優先（因為有完整的 matches 列表）
      2. keyword 結果補充（避免重複）
    """
    merged = []
    seen_row_ids = set()

    def _get_row_id(item: dict) -> str:
        """提取 row_id"""
        rid = item.get("row_id")
        if rid:
            return str(rid).strip()
        # fallback: 用 behavior 前 60 字當 key
        return f"behavior:{_norm_text(item.get('behavior', ''))[:60]}"

    # 先加入 embedding 結果
    for item in embedding_results:
        rid = _get_row_id(item)
        if rid not in seen_row_ids:
            seen_row_ids.add(rid)
            merged.append(item)

    # 再加入 keyword 結果（去重）
    for item in keyword_results:
        rid = _get_row_id(item)
        if rid not in seen_row_ids:
            seen_row_ids.add(rid)
            merged.append(item)

    return merged


def call_rag_semantic_search(question: str, top_k: int = 20, score_threshold: float = 0.5, max_depts: int = 5, database: list = None, multifacet_tag: bool = False):
    """
    呼叫 RAG 語意搜尋 + 智能聚合

    參數:
        question: 問卷題目
        top_k: RAG 返回的候選數量（預設 20）
        score_threshold: 相關性閾值 0-1（預設 0.5）
        max_depts: 單題最多返回的部門數（預設 5）

    回傳格式:
    [
      {
        "article": "條文內容",
        "category": "類別",
        "behavior": "行為準則",
        "best_score": 0.89,  # 最高分
        "matches": [
          {
            "keyword": "供應鏈資訊揭露",
            "dept": "採購",
            "impact": "公司為富采集團...",
            "score": 0.89
          },
          {
            "keyword": "財務資訊揭露",
            "dept": "股務",
            "impact": "公司為富采集團...",
            "score": 0.72
          }
        ]
      }
    ]
    """
    rag_debug = os.getenv("RAG_DEBUG", "0") == "1"

    # ----------------------------
    # ✅ 輕量 must-have gate（非 rerank）
    # ----------------------------
    def _must_have_terms(q: str) -> list:
        """
        輕量過濾：只針對「明確主題」做初步檢測，不做硬性剔除。
        - 童工/未成年 → 記錄標記，由 keyword 補撈來處理
        """
        q2 = (q or "").lower()
        if any(k in q2 for k in ["童工", "未成年", "最低工作年齡", "未成年工", "child labor", "young worker", "underage"]):
            return ["童工", "未成年", "最低工作年齡", "未成年工", "child labor", "young worker", "underage"]
        return []

    def _passes_must_have(q: str, item: dict) -> bool:
        """
        新邏輯：不做硬性剔除，全部通過
        """
        return True

    url = f"{RAG_CONFIG['base_url']}/api/pdf/semantic_query"
    headers = {
        "X-N8N-API-KEY": RAG_CONFIG["api_key"],
        "Content-Type": "application/json",
    }

    print(f"  [RAG] 呼叫語意搜尋 API (top_k={top_k})，題目: {question[:30]}...")

    # 動態取得 PDF ID（必須成功，否則 raise RuntimeError）
    pdf_id = get_latest_pdf_id()

    try:
        payload = {
            "query": question,
            "unit": RAG_CONFIG["unit"],
            "pdf_id": pdf_id,  # 一律帶入，確保只搜尋正確知識庫
        }




        resp = requests.post(url, json=payload, headers=headers, timeout=15)

        # ✅ 檢查 SSO 認證問題
        if resp.status_code == 401:
            error_detail = ""
            try:
                error_json = resp.json()
                error_detail = error_json.get("detail", "")
            except:
                pass

            if "SSO" in error_detail or "登入" in error_detail:
                raise Exception(f"RAG 服務需要 SSO 認證，但服務間認證失敗。請聯繫系統管理員檢查 RAG 服務配置（port 8004）。錯誤詳情：{error_detail}")
            else:
                raise Exception(f"RAG 服務認證失敗（401 Unauthorized）：{error_detail}")

        resp.raise_for_status()
        data = resp.json()

        if not data:
            print("  [RAG] 搜尋結果為空")
            return []

        # 只取前 top_k 個（如果 RAG 不支援 top_k 參數，在這裡手動截斷）
        data = data[:top_k]

        # ========== 按 row_id 聚合（保證 D/E 同列）==========
        article_groups = {}

        for idx, item in enumerate(data):
            # [print-EMB1][上面]：檢查 RAG API 回傳的欄位
            if os.getenv("RAG_SKIP_DEBUG", "0") == "1" and idx == 0:
                raw_keys = sorted(list(item.keys()))[:30]
                print(f"[print-EMB1][上面] rag_item_keys={raw_keys}")

            score = float(item.get("score", 0.0))
            text = item.get("text", "")

            # 解析欄位（row_id 最優先）
            row_id = extract_field(text, "[row_id]")

            # ✅ 強化 row_id 提取：使用多路 fallback
            if not row_id:
                # 嘗試從 item 的其他欄位或 text 中提取
                normalized = normalize_rag_item(item, fallback_prefix="EMB")
                row_id = normalized.get("row_id", "")

            article = extract_field(text, "[條文內容]")
            category = extract_field(text, "[類別]")
            behavior = extract_field(text, "[行為準則]")
            keyword = extract_field(text, "[關鍵字]")
            dept = extract_field(text, "[權責部門]")
            impact = extract_field(text, "[目前現況]")

            # [print-EMB2][下面]：檢查 row_id 提取結果
            if os.getenv("RAG_SKIP_DEBUG", "0") == "1" and idx < 3:
                behavior_preview = (behavior or "")[:30]
                print(f"[print-EMB2][下面] idx={idx} row_id={row_id} score={score:.3f} behavior={behavior_preview}")

            # ✅ 強制同列一致：dept / impact 任何一個抽不到，就不要進來（避免 D/E 拆列）
            if not dept.strip() or not impact.strip():
                # [print-R1] 可抽樣看是哪個 row 抽不到（調試用，可註解）
                # print(f"[print-R1] drop incomplete pair row_id={row_id or 'NA'} score={score:.3f}")
                continue

            # 改用 row_id 分組（不是 article）
            row_key = row_id.strip() if row_id else f"unknown_row_{idx}"

            if row_key not in article_groups:
                article_groups[row_key] = {
                    "row_id": row_id,
                    "article": article,
                    "category": category,
                    "behavior": behavior,
                    "best_score": score,
                    "matches": []
                }

            # best_score 更新
            if score > article_groups[row_key]["best_score"]:
                article_groups[row_key]["best_score"] = score

            # 累積匹配項
            article_groups[row_key]["matches"].append({
                "row_id": row_id,
                "keyword": keyword or "未知",
                "dept": dept,
                "impact": impact,
                "score": score,
            })

        # ========== Keyword 補撈：對所有題目做關鍵字搜尋 ==========
        # 觸發條件：
        #   1. 多面向題（multifacet_tag=True）→ 一律補撈，確保各面向都有 keyword 候選
        #   2. 單面向題，RAG 召回 < 5    → 語意覆蓋不足時補救
        # 設計理念：不以「數量」判斷，而以「是否有機會補足面向缺口」判斷
        kw_candidates = []
        rag_recall_count = len(article_groups)

        if (multifacet_tag or rag_recall_count < 5) and database:
            trigger_reason = "多面向題補全各面向" if multifacet_tag else f"RAG召回不足({rag_recall_count})"
            print(f"  [KW-TRIGGER] {trigger_reason}，使用題目關鍵字補撈")
            
            # 直接用題目拆詞作為搜尋詞
            import jieba
            words = jieba.cut(question)
            search_terms = [w for w in words if len(w) > 1]
            
            # 加入常見關鍵字
            common_keywords = [
                "童工", "未成年", "工時", "加班", "保證金", "證件", "身分證", 
                "護照", "歧視", "種族", "性別", "年齡", "殘疾", "懷孕",
                "結社", "工會", "集會", "自由", "溝通", "主管", "健康", 
                "安全", "環境", "危險", "消防", "急救"
            ]
            for kw in common_keywords:
                if kw in question:
                    search_terms.append(kw)
            
            # 移除重複，限制數量
            search_terms = list(dict.fromkeys(search_terms))[:20]
            
            print("  [KW-TERMS] 搜尋詞: {}".format(search_terms[:10]))
            print("  [KW-DEBUG] database={}".format(type(database).__name__ if database else "None"))
            
            # 執行 keyword 搜尋
            try:
                kw_candidates = _keyword_search_attachment3(database, search_terms, limit=15)
                print("  [KW-RESULT] keyword命中={}".format(len(kw_candidates)))
            except Exception as e:
                print("  [KW-ERROR] {}".format(str(e)))
                kw_candidates = []
            
            # 合併到 article_groups
            for kw_item in kw_candidates:
                row_key = kw_item.get("row_id", "")
                if row_key and row_key not in article_groups:
                    article_groups[row_key] = kw_item
            
            print("  [KW-MERGE] 合併後候選數={}".format(len(article_groups)))

        # ----------------------------
        # ✅ Debug log + must-have gate（在 threshold 過濾之前）
        # ----------------------------
        # 計算原始 embedding 和 keyword 數量
        embedding_count = len([g for g in article_groups.values() if g.get("best_score", 0) > 0])
        keyword_count = len(kw_candidates) if database and 'kw_candidates' in locals() else 0

        if rag_debug:
            # [print-RAG0] 原始召回統計
            print(f"  [print-RAG0][下面] raw_counts embedding={embedding_count} keyword={keyword_count} merged={len(article_groups)}")

            # [print-RAG1] 過濾前 top 分數分布
            all_scores = sorted([float(g.get("best_score", 0.0)) for g in article_groups.values()], reverse=True)
            top5_scores = all_scores[:5]
            print(f"  [print-RAG1][下面] pre_filter top_scores={top5_scores} threshold={score_threshold}")

        # must-have gate：剔除不符合主題的候選
        before_gate = len(article_groups)
        gated_groups = {k: v for k, v in article_groups.items() if _passes_must_have(question, v)}
        removed_by_gate = before_gate - len(gated_groups)

        if rag_debug and removed_by_gate > 0:
            # [print-RAG2] must-have gate 剔除統計
            print(f"  [print-RAG2][下面] must_have_gate removed={removed_by_gate} remain={len(gated_groups)}")

        # 使用過濾後的 gated_groups 進行後續處理
        article_groups = gated_groups

        # ========== 智能過濾每個條文的相關面向 ==========
        results = []
        for article_key, group in article_groups.items():
            matches = group["matches"]

            # 策略 1: 分數閾值過濾
            relevant_matches = [
                m for m in matches
                if m["score"] >= score_threshold
            ]

            # 策略 2: 如果過濾後為空，至少保留 top-1
            if not relevant_matches and matches:
                relevant_matches = [max(matches, key=lambda x: x["score"])]

            # 策略 3: 去重部門（相同部門只保留最高分的）
            dept_best = {}
            for m in relevant_matches:
                dept = m["dept"]
                if dept not in dept_best or m["score"] > dept_best[dept]["score"]:
                    dept_best[dept] = m

            relevant_matches = list(dept_best.values())

            # 策略 4: 限制最多數量
            relevant_matches = sorted(
                relevant_matches,
                key=lambda x: -x["score"]
            )[:max_depts]

            if relevant_matches:
                results.append({
                    "row_id": group.get("row_id", ""),  # ✅ 修正：加入 row_id
                    "article": group["article"],
                    "category": group["category"],
                    "behavior": group["behavior"],
                    "best_score": group["best_score"],
                    "_source": group.get("_source", "semantic"),  # [方案二-B] 保留來源標記
                    "matches": relevant_matches,
                })

        # 按最高分排序
        results.sort(key=lambda x: -x["best_score"])

        # [print-RAG3] 當 threshold 過濾導致結果為 0 時，顯示診斷資訊
        if rag_debug and len(results) == 0 and len(article_groups) > 0:
            # 顯示 gate 之後但 threshold 之前的 top5 分數
            pre_threshold_scores = sorted([float(g.get("best_score", 0.0)) for g in article_groups.values()], reverse=True)[:5]
            print(f"  [print-RAG3][下面] threshold_filtered_to_zero top_scores_after_gate={pre_threshold_scores} threshold={score_threshold}")

        print(f"  [RAG] 聚合後 {len(results)} 個條文")
        for i, r in enumerate(results[:3]):
            print(f"    [{i}] score={r['best_score']:.3f}, 匹配數={len(r['matches'])}, 條文={r['article'][:40] if r['article'] else '無'}...")

        return results

    except Exception as e:
        error_msg = str(e)
        print(f"  [RAG] 呼叫語意搜尋失敗: {repr(e)}")

        # ✅ 如果是 SSO 認證錯誤，拋出特殊異常以便外層處理
        if "SSO" in error_msg or "認證失敗" in error_msg or "401" in error_msg:
            raise Exception(f"RAG_AUTH_ERROR: {error_msg}")

        return []



# =========================================================
# 格式自動偵測：動態判斷問卷欄位對應
# =========================================================
import string as _string

def _detect_format_with_llm(llm_cfg, sample_text, max_cols):
    """使用 LLM 分析問卷欄位格式"""
    url = f"{llm_cfg.endpoint}/openai/deployments/{llm_cfg.deployment}/chat/completions"
    headers = {"api-key": llm_cfg.api_key, "Content-Type": "application/json"}

    system_prompt = "你是 Excel 問卷格式分析專家。分析以下工作表前幾行，判斷各欄位的角色。只回傳 JSON，不要有其他文字。"
    user_prompt = f"""以下是 Excel 工作表的前幾行（格式：Row N: [欄A, 欄B, ...]）：

{sample_text}

回傳 JSON（只回 JSON）：
{{
  "start_row": <資料起始行整數，跳過標題>,
  "clause_col": "<條款/序號欄字母，如 A，若無則 null>",
  "question_col": "<題目/條文/要求主要內容欄字母，必填>",
  "dept_col": "<填寫權責部門欄字母，若無則 null>",
  "impact_col": "<填寫現況/影響欄字母，若無則 null>",
  "reasoning": "<說明，30字內>"
}}

規則：
- clause_col：值是短編號（1.1、A.2、勞工標準）→ 條款欄
- question_col：最長、最主要的文字說明/要求/問題欄（必填）
- dept_col：標題含「部門」「權責」「負責」
- impact_col：標題含「現況」「影響」「說明」「回答」「鑑定」
- 若分不清 clause 與 question，優先選最長文字欄為 question_col"""

    payload = {
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
        "max_completion_tokens": 2000,  # P0 Fix: gpt-5.3-chat 推理型模型用 max_completion_tokens
    }
    resp = requests.post(url, headers=headers, json=payload, params={"api-version": AZURE_OPENAI_API_VERSION}, timeout=20)
    resp.raise_for_status()

    raw = resp.json()["choices"][0]["message"]["content"].strip()
    raw = re.sub(r"```json\s*", "", raw)
    raw = re.sub(r"```\s*", "", raw)

    data = json.loads(raw)
    return {
        "start_row": int(data.get("start_row", 3)),
        "clause_col": (data.get("clause_col") or None),
        "question_col": str(data.get("question_col", "C")).upper(),
        "dept_col": (data.get("dept_col") or None),
        "impact_col": (data.get("impact_col") or None),
    }


def _detect_format_with_keywords(ws, max_cols):
    """關鍵字 fallback：掃標題行推斷欄位"""
    # 找標題行（非空欄最多的那行，掃前5行）
    header_row = 1
    header_vals = {}
    for r in range(1, min(6, ws.max_row + 1)):
        row_vals = {}
        for c in range(1, max_cols + 1):
            col_letter = _string.ascii_uppercase[c - 1]
            val = ws[f"{col_letter}{r}"].value
            if val and str(val).strip():
                row_vals[col_letter] = str(val).strip().lower()
        if len(row_vals) > len(header_vals):
            header_vals = row_vals
            header_row = r

    clause_kw  = ["條款", "序號", "item", "clause", "no.", "編號", "article"]
    question_kw= ["題目", "條文", "問題", "要求", "內容", "question", "requirement", "description", "承諾"]
    dept_kw    = ["部門", "權責", "負責", "department", "owner"]
    impact_kw  = ["現況", "影響", "說明", "回答", "鑑定", "備註", "status", "comment", "response", "鑑別"]

    clause_col = question_col = dept_col = impact_col = None

    for col, val in header_vals.items():
        if any(kw in val for kw in clause_kw) and not clause_col:
            clause_col = col
        elif any(kw in val for kw in question_kw) and not question_col:
            question_col = col
        elif any(kw in val for kw in dept_kw) and not dept_col:
            dept_col = col
        elif any(kw in val for kw in impact_kw) and not impact_col:
            impact_col = col

    # 若 question_col 仍找不到，取樣本中平均長度最長的欄
    if not question_col:
        best_col, best_len = None, 0
        for c in range(1, max_cols + 1):
            col_letter = _string.ascii_uppercase[c - 1]
            if col_letter in {clause_col, dept_col, impact_col}:
                continue
            total, count = 0, 0
            for r in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
                v = ws[f"{col_letter}{r}"].value
                if v:
                    total += len(str(v)); count += 1
            avg = total / count if count else 0
            if avg > best_len:
                best_len, best_col = avg, col_letter
        question_col = best_col or "C"

    return {
        "start_row": header_row + 1,
        "clause_col": clause_col,
        "question_col": question_col,
        "dept_col": dept_col,
        "impact_col": impact_col,
    }


def detect_survey_format(ws, llm_cfg=None):
    """
    動態偵測問卷 Excel 的欄位對應。
    LLM 優先，失敗則 fallback 到關鍵字偵測。
    返回: {start_row, clause_col, question_col, dept_col, impact_col}
    """
    max_cols = min(ws.max_column, 10)

    # 取前12行做樣本
    sample_rows = []
    for r in range(1, min(13, ws.max_row + 1)):
        row_data = []
        for c in range(1, max_cols + 1):
            col_letter = _string.ascii_uppercase[c - 1]
            val = ws[f"{col_letter}{r}"].value
            row_data.append(str(val)[:100] if val else "")
        sample_rows.append(f"Row {r}: {row_data}")
    sample_text = "\n".join(sample_rows)

    if llm_cfg and _HAS_LLM_FIRST:
        try:
            result = _detect_format_with_llm(llm_cfg, sample_text, max_cols)
            if result and result.get("question_col"):
                print(f"[格式偵測][LLM] start={result['start_row']} clause={result['clause_col']} question={result['question_col']} dept={result['dept_col']} impact={result['impact_col']}")
                return result
        except Exception as e:
            print(f"[格式偵測][LLM] 失敗: {e}，改用 keyword fallback")

    result = _detect_format_with_keywords(ws, max_cols)
    print(f"[格式偵測][Keyword] start={result['start_row']} clause={result['clause_col']} question={result['question_col']} dept={result['dept_col']} impact={result['impact_col']}")
    return result


def process_workbook(survey_path, db_path, client_name="未指定客戶"):
    """
    處理問卷 Excel 的主函數

    步驟：
      1. 載入附件2（問卷）和附件3（資料庫）
      2. 逐題執行 two_stage_matching
      3. 寫回 D 欄（權責部門）和 E 欄（現況/影響）
      4. 產生比對報告

    回傳：
      {
        "output_path": 輸出檔案路徑,
        "report": {
          "total_questions": 總題數,
          "matched_count": 成功匹配數,
          "low_confidence_count": 低信心題目數,
          "details": [每題詳細資訊]
        }
      }
    """

    print("\n" + "=" * 80)
    print(f"[問卷處理] 開始處理問卷")
    print("=" * 80)


    # ========== 初始化 LLM-first 配置 ==========
    llm_cfg = None
    if _HAS_LLM_FIRST:
        try:
            llm_cfg = AzureOpenAIConfig(
                endpoint=AZURE_OPENAI_ENDPOINT,
                api_key=AZURE_OPENAI_API_KEY,
                deployment=AZURE_OPENAI_DEPLOYMENT,
                api_version=AZURE_OPENAI_API_VERSION,
            )
            print(f"[LLM-first] ✅ 配置成功: {AZURE_OPENAI_ENDPOINT}/{AZURE_OPENAI_DEPLOYMENT}")
        except Exception as e:
            print(f"[LLM-first] ⚠️  配置失敗: {e}")
            llm_cfg = None
    else:
        print("[LLM-first] ⚠️  模組未載入，將使用傳統比對")

    # ========== 1. 載入資料庫 ==========
    database = load_database(db_path)

    # ========== 2. 載入問卷 Excel ==========
    print(f"\n[問卷載入] 開始讀取: {survey_path}")

    # ★ 保留外部連結以維持結構完整性，避免 Excel 修復警告
    wb = openpyxl.load_workbook(survey_path, keep_links=True)
    ws = wb.active

    print(f"[問卷載入] 工作表名稱: {ws.title}")
    print(f"[問卷載入] 最大行數: {ws.max_row}")

    # ========== 2.5 格式自動偵測 ==========
    fmt = detect_survey_format(ws, llm_cfg)
    start_row    = fmt['start_row']
    clause_col   = fmt['clause_col']         # 可能是 None（無條款欄）
    question_col = fmt['question_col']
    dept_col     = fmt['dept_col']   or 'D'  # fallback
    impact_col   = fmt['impact_col'] or 'E'  # fallback

    # ========== 3. 逐題比對 ==========
    results = []
    matched_count = 0
    low_confidence_count = 0

    # ✅ 追蹤 RAG 認證失敗：如果前 5 題都失敗且是 SSO 問題，立即中斷
    rag_auth_error_count = 0
    rag_auth_error_msg = None

    print(f"\n[比對開始] 準備處理題目...")

    # start_row 已由 detect_survey_format 決定

    # ====== 建立「條款編號(B欄) → row_idx」索引（題目順序不固定也能跳）======
    def _norm_clause(x) -> str:
        return str(x).strip().upper() if x is not None else ""

    clause_to_row = {}
    if clause_col:
        for r in range(start_row, ws.max_row + 1):
            c = _norm_clause(ws[f"{clause_col}{r}"].value)
            if c:
                clause_to_row[c] = r

    # ====== SKIP controller ======
    skip_until_clause = ""   # 例如 "C.7.1"
    skip_reason = ""

    # [print-SK0] ✅ 加在「下面」
    print(f"[print-SK0][下面] clause index built={len(clause_to_row)}")

    for row_idx in range(start_row, ws.max_row + 1):
        clause = _norm_clause(ws[f"{clause_col}{row_idx}"].value) if clause_col else ""

        # [print-SK1] ✅ 加在「下面」
        if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
            print(f"[print-SK1][下面] row={row_idx} clause={clause} skip_until={skip_until_clause}")

        # 若在 skip 狀態，且尚未到目標條款 → 直接填 SKIP 並跳過
        if skip_until_clause and clause and clause != skip_until_clause:
            ws[f"{dept_col}{row_idx}"] = clean_excel_text(f"SKIP（依 {skip_reason}）")
            ws[f"{impact_col}{row_idx}"] = clean_excel_text(f"SKIP（依 {skip_reason}）")
            ws[f"{dept_col}{row_idx}"].alignment = ALIGN_D_COLUMN
            ws[f"{impact_col}{row_idx}"].alignment = ALIGN_E_COLUMN
            ws[f"{dept_col}{row_idx}"].fill = FILL_YELLOW
            ws[f"{impact_col}{row_idx}"].fill = FILL_YELLOW
            continue

        # 到了目標條款 → 解除 skip，繼續正常作答
        if skip_until_clause and clause == skip_until_clause:
            skip_until_clause = ""
            skip_reason = ""

        # 讀取 A 欄類別和 C 欄題目
        category = ws[f"A{row_idx}"].value
        category = str(category).strip() if category else ""

        question_cell = ws[f"{question_col}{row_idx}"]
        question = question_cell.value

        # 跳過空白題目
        if not question or not str(question).strip():
            continue

        question = str(question).strip()

        # ========== ✅ 法律條款偵測（違約罰則/合約效力）==========
        if is_legal_clause(question):
            print(f"[法律條款] row={row_idx} 識別為法律條款，標記黃色")
            dept_text = "違約罰則、合約效力等法律條款，請依人工比對"
            impact_text = "此條款屬合約法律性質，知識庫無對應RBA行為準則，請人工填寫"
            ws[f"{dept_col}{row_idx}"] = clean_excel_text(dept_text)
            ws[f"{impact_col}{row_idx}"] = clean_excel_text(impact_text)
            ws[f"{dept_col}{row_idx}"].alignment = ALIGN_D_COLUMN
            ws[f"{impact_col}{row_idx}"].alignment = ALIGN_E_COLUMN
            ws[f"{dept_col}{row_idx}"].fill = FILL_YELLOW
            ws[f"{impact_col}{row_idx}"].fill = FILL_YELLOW
            results.append({
                "row": row_idx,
                "question": question[:50] + "...",
                "matched": False,
                "confidence": 0,
                "method": "legal_clause",
                "dept": dept_text,
                "impact": impact_text
            })
            continue  # 不計入 matched_count

        # ========== ✅ 承諾聲明檢查（免 RAG） ==========
        if is_commitment_statement(question):
            if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
                print(f"[Commitment] row={row_idx} 識別為承諾聲明，標記為 -")
            
            ws[f"{dept_col}{row_idx}"] = clean_excel_text("-")
            ws[f"{impact_col}{row_idx}"] = clean_excel_text("-")
            ws[f"{dept_col}{row_idx}"].alignment = ALIGN_D_COLUMN
            ws[f"{impact_col}{row_idx}"].alignment = ALIGN_E_COLUMN
            ws[f"{dept_col}{row_idx}"].fill = FILL_GREEN
            ws[f"{impact_col}{row_idx}"].fill = FILL_GREEN
            
            results.append({
                "row": row_idx,
                "question": question[:50] + "...",
                "matched": True,
                "confidence": 100,
                "method": "commitment_statement",
                "dept": "-",
                "impact": "-"
            })
            matched_count += 1
            continue

        # ========== ✅ LLM-first 萃取題目 Schema ==========
        q_schema = None
        if llm_cfg and _HAS_LLM_FIRST:
            try:
                if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
                    print(f"[LLM-first][上面] row={row_idx} 開始萃取...")
                
                q_schema = extract_question_schema_llm(
                    cfg=llm_cfg,
                    clause_id=clause or "",
                    category=category,
                    question_text=question,
                    debug=(os.getenv("RAG_SKIP_DEBUG", "0") == "1"),
                )
                
                if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
                    print(f"[LLM-first][下面] intent={q_schema.intent[:50]}...")
                    print(f"[LLM-first][下面] must_have={q_schema.must_have_terms[:5]}")
                    print(f"[LLM-first][下面] topics={q_schema.topic_tags[:3]}")
                    print(f"[LLM-first][下面] confidence={q_schema.confidence:.2f}")
                
            except Exception as e:
                print(f"[LLM-first] ⚠️  row={row_idx} 萃取失敗: {e}")
                q_schema = None


        # ★ print-S1：跳過前言/說明文字
        if question in ("前言", "(前言)"):
            print(f"\n[題目 {row_idx}] 偵測到前言/說明文字，跳過比對: {question}")
            continue

        print(f"\n[題目 {row_idx}] {question[:50]}...")

        # ✅ 初始化變數（避免在某些分支中未定義導致 UnboundLocalError）
        no_skip_target = ""

        # ✅ [print-RQ1] 加在「下面」：顯示查詢補詞後的 query（避免你以為 rerank 沒效，其實是沒召回）
        # ========== ✅ 組裝 RAG 查詢字串（LLM-first 或傳統補詞） ==========
        if q_schema and q_schema.confidence > 0.3:
            # 用 LLM-first schema 組裝查詢
            rag_query = f"{q_schema.intent} {' '.join(q_schema.must_have_terms)} {' '.join(['#' + tag for tag in q_schema.topic_tags])}"
            if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
                print(f"[查詢][下面] LLM-first Query: {rag_query[:120]}...")
        else:
            # Fallback：傳統補詞
            rag_query = expand_question_for_rag(question)
            if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
                print(f"[查詢][下面] 傳統補詞 Query: {rag_query[:120]}...")

        if os.getenv("RAG_SKIP_DEBUG", "0") == "1" and rag_query != question:
            print(f"[print-RQ1][下面] expanded_query='{rag_query[:160]}...'")

        # ✅ [print-RAGCFG] 加在「下面」：確認 threshold/top_k（避免你以為沒召回，其實是被 threshold 擋掉）
        if os.getenv("RAG_DEBUG", "0") == "1":
            print(f"[print-RAGCFG][下面] top_k=20 score_threshold=0.50 max_depts=5")

        # ========== 改用 RAG 語意搜尋 (top-k + 聚合 + Hybrid 召回) ==========
        # 先預計算 multifacet_tag，讓 call_rag_semantic_search 決定是否補撈
        # （正式計算在後面 if candidates 區塊，此處只做 early detection）
        _pre_multifacet_tag = is_multifacet_question_expected(question, q_schema=q_schema)
        try:
            candidates = call_rag_semantic_search(
                rag_query,             # ✅ 改：用補詞後 query 送 RAG
                top_k=20,              # 取 top-20
                score_threshold=0.5,  # 閾值
                max_depts=5,           # 最多 5 個部門
                database=database,     # ✅ 傳入附件三 database 做 keyword 補撈
                multifacet_tag=_pre_multifacet_tag  # ✅ 多面向題一律補撈
            )
        except Exception as e:
            error_msg = str(e)

            # ✅ 檢測 RAG 認證錯誤
            if "RAG_AUTH_ERROR" in error_msg:
                rag_auth_error_count += 1
                rag_auth_error_msg = error_msg.replace("RAG_AUTH_ERROR: ", "")

                print(f"  ⚠️  偵測到 RAG 認證錯誤（第 {rag_auth_error_count} 次）")

                # 如果連續 5 題都是認證錯誤，立即中斷
                if rag_auth_error_count >= 5:
                    print(f"\n{'=' * 80}")
                    print(f"[錯誤] RAG 服務認證持續失敗，已中斷處理")
                    print(f"[錯誤] 錯誤訊息：{rag_auth_error_msg}")
                    print(f"{'=' * 80}")
                    wb.close()
                    raise Exception(f"RAG 服務認證失敗，無法繼續處理問卷。{rag_auth_error_msg}")

                candidates = []
            else:
                # 其他錯誤，重新拋出
                raise

        # ✅ [題域閘門] 在進入候選處理前，過濾掉跨題域的候選
        if candidates and _HAS_FACET_TAXONOMY:
            _taxonomy_entry = get_taxonomy_entry(clause or "", question)
            if _taxonomy_entry:
                _before = len(candidates)
                candidates = domain_gate_filter(
                    candidates,
                    allowed_categories=_taxonomy_entry.get("allowed_categories", []),
                    excluded_kw=_taxonomy_entry.get("excluded_kw", []),
                )
                print(f"  [DomainGate] {_before}→{len(candidates)} 個候選（條款:{clause}）")

        if candidates:

            # ========== ✅ LLM-first 優化：Coverage Gating + Exact Token Boost ==========
            if q_schema and len(candidates) > 0 and ENABLE_RERANK:
                if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
                    print(f"[Optimize][上面] 開始優化 {len(candidates)} 個候選條文...")

                # 1. Coverage Gating（多概念覆蓋度評分）
                if q_schema.must_have_terms and len(q_schema.must_have_terms) > 1:
                    candidates = rerank_with_coverage_gating(
                        candidates,
                        q_schema.must_have_terms,
                        debug=(os.getenv("RAG_SKIP_DEBUG", "0") == "1")
                    )

                # 2. Exact Token Boost（精確數字/標準匹配）
                candidates = exact_token_boost(
                    candidates,
                    q_schema,
                    debug=(os.getenv("RAG_SKIP_DEBUG", "0") == "1")
                )

                if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
                    print(f"[Optimize][下面] 優化完成，top-3 分數:")
                    for i, cand in enumerate(candidates[:3], 1):
                        print(f"  [{i}] {cand.get('behavior', '')[:30]} | score={cand.get('best_score', 0):.1f}")


            # ✅ 預先計算 facets（供後續 per-facet 選法使用）
            # 優先使用受控 facet 白名單（若條款在 taxonomy 中）
            _taxonomy_entry_facet = get_taxonomy_entry(clause or "", question) if _HAS_FACET_TAXONOMY else None
            if _taxonomy_entry_facet and llm_cfg:
                # 受控路徑：LLM 從白名單選 facet（防漂移）
                _facet_ids = get_controlled_facets_cached(llm_cfg, clause or "", question, _taxonomy_entry_facet)
                facets_expected = [
                    f["zh"] for f in _taxonomy_entry_facet.get("canonical_facets", [])
                    if f["id"] in _facet_ids
                ]
                if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
                    print(f"  [FacetCtrl] 受控選取: {_facet_ids} → {facets_expected}")
            else:
                # Fallback：原有自由生成邏輯
                facets_expected = get_facets_expected_from_question(question, q_schema=q_schema)
            multifacet_tag = len(facets_expected) >= 2

            # ✅ 改：不要直接用 top-1，先 rerank；若 rerank 判定「候選缺核心詞」→ 回 None → 視為未匹配
            if ENABLE_RERANK:
                selected = rerank_rag_candidates(question, candidates, top_n=8)

                if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
                    # [print-RR2] 加在「下面」：rerank 是否改變 top-1 / 或直接拒答
                    top1_behavior = (candidates[0].get("behavior") or "")[:40]
                    best_behavior = ((selected or {}).get("behavior") or "")[:40]
                    changed = "REJECT" if selected is None else ("YES" if top1_behavior != best_behavior else "NO")
                    print(f"[print-RR2][下面] rerank_changed={changed} top1='{top1_behavior}...' best='{best_behavior}...'")
            else:
                # ENABLE_RERANK = False: top-1→5 pre-selection fallback
                # 依序嘗試 top-1~5，選第一個能通過 keyword_verify 的候選
                _q_for_prescan = (q_schema.question_text_zh if q_schema and q_schema.question_text_zh else question)
                selected = None
                for _pi, _pc in enumerate(candidates[:5]):
                    _pct = " ".join([
                        _pc.get("article") or "",
                        _pc.get("behavior") or "",
                        " ".join((m.get("keyword") or "") for m in _pc.get("matches", [])),
                        " ".join((m.get("dept") or "") for m in _pc.get("matches", [])),
                    ])
                    if verify_keyword_match(_q_for_prescan, _pct, min_match=1):
                        selected = _pc
                        if _pi > 0:
                            print(f"  [KW-FALLBACK] top-1 keyword預選失敗，改用第{_pi+1}個候選({(_pc.get('behavior') or '')[:20]})")
                        break
                    elif _pi == 0:
                        # [Top1Trust-KW] top-1 KW 弱匹配，但 RAG 高分(≥0.60)時仍信任
                        # 防止 KW-FALLBACK 在 DomainGate 已確認分類後仍替換正確答案（如 Row56 文件和紀錄）
                        _pc_kw_score = float(_pc.get("best_score", 0.0))
                        if _pc_kw_score >= 0.60:
                            selected = _pc
                            print(f"  [Top1Trust-KW] top-1 KW弱匹配但score={_pc_kw_score:.3f}≥0.60，信任RAG top-1: {(_pc.get('behavior') or '')[:30]}")
                            break
                if selected is None:
                    selected = candidates[0] if candidates else None  # 全部失敗→交由後續 keyword_verify 處理
                if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
                    print(f"[print-RR2][下面] ENABLE_RERANK=False，top-1→5 pre-scan 完成")

            if selected is None:
                candidates = []  # ✅ 讓後面走「未匹配」分支（像人工：找不到就不硬答）
            else:
                # ====== 多面向題：per-facet 選法（不影響單面向邏輯） ======
                # [Fix-Top1Trust] RAG top-1 明顯主導時跳過 per-facet，防止偏移 facet 替換正確答案
                # 例：Row56 文件和紀錄 top-1=0.575, top-2=0.518, gap=0.057 -> 直接信任 top-1
                _top1_score = float(candidates[0].get("best_score", 0.0)) if candidates else 0.0
                _top2_score = float(candidates[1].get("best_score", 0.0)) if len(candidates) > 1 else 0.0
                _top1_dominant = (_top1_score - _top2_score) >= 0.05  # gap >= 5% 視為主導
                if multifacet_tag and len(facets_expected) >= 2 and not _top1_dominant:
                    per_facet_list = select_per_facet_candidates(candidates, facets_expected)

                    # ✅ [Per-Facet 驗證] 過濾掉不符合 required_kw 的候選（防止錯誤 facet 混入）
                    if _HAS_FACET_TAXONOMY and _taxonomy_entry_facet:
                        _canonical_map = {
                            f["zh"]: f for f in _taxonomy_entry_facet.get("canonical_facets", [])
                        }
                        _validated_list = []
                        for _cand in per_facet_list:
                            # 找這個候選最可能對應的 facet_def
                            _best_facet_def = None
                            _best_kw_score = -1
                            for _fterm, _fdef in _canonical_map.items():
                                _hits = sum(1 for kw in _fdef.get("required_kw", [])
                                            if kw in " ".join([
                                                _cand.get("behavior", ""),
                                                _cand.get("article", ""),
                                            ]))
                                if _hits > _best_kw_score:
                                    _best_kw_score, _best_facet_def = _hits, _fdef
                            is_valid, reason = validate_facet_candidate(_cand, _best_facet_def or {})
                            if is_valid:
                                _validated_list.append(_cand)
                            else:
                                print(f"  [FacetValidate] ❌ 淘汰無效候選: {_cand.get('behavior','')[:30]} ({reason})")
                        if _validated_list:
                            per_facet_list = _validated_list
                        elif per_facet_list:
                            print(f"  [FacetValidate] ⚠️  所有候選驗證失敗，保留原始清單（安全網）")

                    if len(per_facet_list) > 1:
                        # 合併多組 matches（各 facet 最佳候選的 dept/impact 全部納入）
                        matches = [m for c in per_facet_list for m in c["matches"]]
                        behavior_parts = list(dict.fromkeys(
                            c.get("behavior", "") for c in per_facet_list if c.get("behavior")))
                        behavior = behavior_parts[0] if len(set(behavior_parts)) == 1 \
                            else "、".join(behavior_parts[:3])
                        confidence = max(
                            float(c.get("best_score", 0.0)) for c in per_facet_list) * 100.0
                        article = per_facet_list[0].get("article") or ""
                        # [多面向增強] keyword_verify 用的擴充文字（含所有 facet 的 article）
                        _multifacet_extra = " ".join(
                            c.get("article","") for c in per_facet_list[1:]
                        )
                        # selected 仍指向主要候選（供後續 LLM upgrade 取 category）
                        selected = max(per_facet_list,
                                       key=lambda c: float(c.get("best_score", 0.0)))
                        print(f"  [多面向選法] per_facet={len(per_facet_list)} 個，"
                              f"合併 {len(matches)} 個 matches")
                    else:
                        # per_facet 命中 1 個：優先用它（比 KW-FALLBACK top-1 更精準）
                        # per_facet 命中 0 個：退回 KW-FALLBACK selected（安全網）
                        if per_facet_list:
                            # ✅ facet 命中候選優先，覆蓋 KW-FALLBACK 可能的弱匹配
                            selected = per_facet_list[0]
                            matches = selected["matches"]
                            behavior = selected.get("behavior") or "無"
                            confidence = float(selected.get("best_score", 0.0)) * 100.0
                            article = selected.get("article") or ""
                            print(f"  [CONF-DEBUG][per-facet use-1] _source={selected.get('_source', '?')!r} raw_score={selected.get('best_score',0):.4f} raw_conf={confidence:.1f}% (facet命中1個，優先使用)")
                        else:
                            # per_facet 完全沒命中 → 退回 KW-FALLBACK selected（安全網）
                            matches = selected["matches"]
                            behavior = selected.get("behavior") or "無"
                            confidence = float(selected.get("best_score", 0.0)) * 100.0
                            article = selected.get("article") or ""
                            print(f"  [CONF-DEBUG][per-facet fallback] _source={selected.get('_source', '?')!r} raw_score={selected.get('best_score',0):.4f} raw_conf={confidence:.1f}% (per_facet無命中，退回KW-FALLBACK)")
                else:
                    # top1_dominant 時覆蓋 selected 為語意最高分候選
                    _t1d = locals().get("_top1_dominant", False)
                    if _t1d and candidates and selected is not candidates[0]:
                        selected = candidates[0]
                        print(f"  [Top1Trust] RAG top-1 主導 (gap={_top1_score - _top2_score:.3f})，覆蓋為: {(selected.get('behavior') or '')[:30]}")
                    matches = selected["matches"]  # ⬅ 單面向 or top1_dominant
                    behavior = selected.get("behavior") or "無"
                    confidence = float(selected.get("best_score", 0.0)) * 100.0
                    article = selected.get("article") or ""
                    print(f"  [CONF-DEBUG][single] _source={selected.get('_source', '?')!r} raw_score={selected.get('best_score',0):.4f} raw_conf={confidence:.1f}%")

                # [Fix-SameBehaviorDept] 同行為準則多部門補強
                # 解決 DB 刻意設計雙部門（如員關+稽核室）被 top-1 遺漏的問題
                # 掃描 candidates，補充與 selected 同 behavior 但不同部門的候選
                _sel_beh = (selected.get("behavior", "") or "").strip() if selected else ""
                if _sel_beh and candidates:
                    _exist_depts = {m.get("dept", "") for m in matches if m.get("dept")}
                    _sel_rid = (selected.get("row_id") or str(id(selected))) if selected else ""
                    for _cand in candidates:
                        if ((_cand.get("behavior", "") or "").strip() == _sel_beh):
                            _crid = _cand.get("row_id") or str(id(_cand))
                            if _crid == _sel_rid:
                                continue  # 跳過 selected 本身（已有）
                            for _m in _cand.get("matches", []):
                                _d = (_m.get("dept") or "").strip()
                                if _d and _d not in _exist_depts:
                                    matches.append(_m)
                                    _exist_depts.add(_d)
                                    print(f"  [SameBehaviorDept] 補充共管部門: {_d} (同行為: {_sel_beh[:25]})")

                # ✅ RAG 匹配結果 log
                if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
                    print(f"[RAG 匹配][下面] 行為準則: {behavior}")
                    print(f"[RAG 匹配][下面] 相似度: {confidence:.1f}")
                    print(f"[RAG 匹配][下面] 匹配數: {len(matches)}")

                print(f"  ✓ RAG 匹配成功 | 相似度分數: {confidence:.1f}")
                print(f"    - 行為準則: {behavior}")
                print(f"    - 匹配數量: {len(matches)}")

                # ====== 先解析「題目是否有 NO 分支跳題」======
                no_skip_target = parse_no_skip_target(question)  # e.g. "C.7.1" or ""
                answer_type = "UNKNOWN"

                # -----------------------------
                # ✅ NEW: Expected(facets) vs Observed(topics) 指標（不影響原流程）
                # -----------------------------
                topk_candidates = candidates[:]  # 聚合後 top-k 候選
                # facets_expected / multifacet_tag 已在 reranker 前計算，直接使用
                topics_observed_dist = compute_topics_observed_dist(topk_candidates, vocab=facets_expected)
                hit_n, total_n = compute_facet_coverage_expected_vs_observed(facets_expected, topics_observed_dist)
                coverage_ratio = f"{hit_n}/{total_n}" if total_n > 0 else "-"
                topic_dist = topics_observed_dist  # 舊鍵名相容（語意=Observed）

                # [print-YR0] ✅ 加在「下面」
                if os.getenv("RAG_YELLOW_DEBUG", "0") == "1":
                    print(f"[print-YR0][下面] multifacet_tag={multifacet_tag} facets_expected={facets_expected} topic_dist={topic_dist} coverage_ratio={coverage_ratio}")

                # 常態多面向診斷輸出（不需 debug flag）
                if multifacet_tag:
                    print(f"  [多面向] facets={facets_expected} coverage={coverage_ratio} topics={list(topics_observed_dist.keys())[:6]}")

                # ====== 關鍵字驗證：避免 plastic 被配到 機密資訊 這種離譜 top-1 ======
                # 用 article + behavior + matches keywords 共同形成 candidate_text（更穩）
                # 多面向：加入其他 facet 的 article 提供更豐富的 keyword 上下文
                _mf_extra = locals().get("_multifacet_extra", "")
                candidate_text = " ".join([
                    article,
                    _mf_extra,
                    behavior,
                    " ".join([(m.get("keyword") or "") for m in matches]),
                    " ".join([(m.get("dept") or "") for m in matches]),
                ])

                # ✅ 優先用繁體版問題文字做關鍵字驗證（避免簡體vs繁體字符不同而誤判未匹配）
                _q_for_verify = (q_schema.question_text_zh if q_schema and q_schema.question_text_zh else question)
                if not verify_keyword_match(_q_for_verify, candidate_text, min_match=1):
                    # keyword mismatch：標記未匹配（紅色）
                    ws[f"{dept_col}{row_idx}"] = clean_excel_text("未匹配（請人工判定）")
                    ws[f"{impact_col}{row_idx}"] = clean_excel_text("知識庫無相關資料或語意不匹配，請人工填寫")
                    ws[f"{dept_col}{row_idx}"].alignment = ALIGN_D_COLUMN
                    ws[f"{impact_col}{row_idx}"].alignment = ALIGN_E_COLUMN
                    ws[f"{dept_col}{row_idx}"].fill = FILL_RED
                    ws[f"{impact_col}{row_idx}"].fill = FILL_RED

                    # [print-KV2] ✅ 加在「下面」
                    print(f"[print-KV2][下面] keyword mismatch → treat as unmatched. no_skip_target={no_skip_target}")

                    low_confidence_count += 1
                    results.append({
                        "row_index": row_idx,
                        "question": question,
                        "matched": False,
                        "confidence": 0,
                        "department": "未匹配（請人工判定）",
                    })

                    # ★ 關鍵：是否觸發跳題，完全由「題目是否存在 NO 跳題指令」決定
                    # 若題目有 NO 分支跳題：為避免後續子題亂答，保守視為 NO → 啟動 skip
                    if no_skip_target:
                        answer_type = "NO"
                        skip_until_clause = no_skip_target
                        skip_reason = f"{clause or row_idx} 未匹配（視為 NO）→ 跳至 {no_skip_target}"
                        # [print-SK2] ✅ 加在「下面」
                        print(f"[print-SK2][下面] enable skip: {skip_reason}")

                    continue  # 跳過後續正常填寫流程

                # ✅ keyword match → 走你原本的 dept_lines/impact_lines 合併與三色判定
                # 後面仍可用 safe_status 再判 YES/NO（如果你需要）

                # ========== 合併多個部門（用換行分隔）==========
                dept_lines = []
                _seen_dept = set()
                for m in matches:
                    # ✅ 只輸出部門，不加關鍵字（避免格式不一致）；去重保序（Row32修復）
                    d = m['dept']
                    if d and d not in _seen_dept:
                        dept_lines.append(d)
                        _seen_dept.add(d)

                department_text = "\n".join(dept_lines)

                # ========== 合併多個現況（用分隔線區分）==========
                impact_lines = []
                for i, m in enumerate(matches):
                    if i > 0:
                        # 分隔線
                        impact_lines.append("─" * 40)

                    # 標註關鍵字
                    if m['keyword'] and m['keyword'] != "未知":
                        impact_lines.append(f"【{m['keyword']}】")

                    # 現況內容
                    impact_lines.append(m['impact'])

                impact_text = "\n".join(impact_lines)

                # ========== 安全清洗 ==========
                safe_department = clean_excel_text(department_text)
                safe_status = clean_excel_text(impact_text)

                # ========== 寫回 D 欄（權責部門）==========
                ws[f"{dept_col}{row_idx}"] = safe_department
                ws[f"{dept_col}{row_idx}"].alignment = ALIGN_D_COLUMN  # 自動換行

                # ========== 寫回 E 欄（現況/影響）==========
                ws[f"{impact_col}{row_idx}"] = safe_status
                ws[f"{impact_col}{row_idx}"].alignment = ALIGN_E_COLUMN  # 自動換行

                # [方案二-C] keyword補撈候選信心補正：通過 keyword_verify 但 score 低
                # semantic 新模型分數 0.07~0.15 映射 confidence 7~15%，遠低於 55 門檻
                # keyword 候選若找到正確條文，應給機會進入黃色 LLM 審查
                _src = selected.get("_source", "semantic") if selected else "semantic"
                if selected and _src == "keyword_local" and confidence < 55.0:
                    print(f"  [KW-CONF-BOOST] keyword補撈候選，信心補正 {confidence:.1f}% → 55.0%（進入黃色 LLM 審查）")
                    confidence = 55.0
                elif selected and _src != "keyword_local" and confidence < 55.0:
                    # [方案一] semantic score calibration：從 MATCH_CONFIG 讀參數
                    # 換模型時只需改 MATCH_CONFIG["sem_score_min/max"]，此處不動
                    _raw_score = float(selected.get("best_score", 0.0))
                    MIN_SEM = MATCH_CONFIG.get("sem_score_min", 0.07)
                    MAX_SEM = MATCH_CONFIG.get("sem_score_max", 0.15)
                    CONF_MIN = MATCH_CONFIG.get("sem_conf_min", 55.0)
                    CONF_MAX = MATCH_CONFIG.get("sem_conf_max", 95.0)
                    _norm = (_raw_score - MIN_SEM) / (MAX_SEM - MIN_SEM)
                    _norm = max(0.0, min(1.0, _norm))  # clamp 防爆出區間
                    _sem_conf = CONF_MIN + _norm * (CONF_MAX - CONF_MIN)
                    print(f"  [SEM-CONF-NORM] semantic候選 _source={_src!r}，"
                          f"raw_score={_raw_score:.4f} raw_conf={confidence:.1f}% "
                          f"→ normalized={_sem_conf:.1f}% "
                          f"(sem_score=[{MIN_SEM},{MAX_SEM}] conf=[{CONF_MIN},{CONF_MAX}])")
                    confidence = _sem_conf

                print(f"  [CONF-FINAL] _source={_src!r} confidence={confidence:.1f}% low_th={MATCH_CONFIG['low_confidence_threshold']} high_th={MATCH_CONFIG.get('high_confidence_threshold',70)}")

                # ========== 三色判定 ==========
                high_th = MATCH_CONFIG.get("high_confidence_threshold", 70)
                low_th = MATCH_CONFIG["low_confidence_threshold"]

                fill = None

                # -----------------------------
                # ✅ NEW: 只在黃/低綠不穩才判斷 yellow_reason（Expected vs Observed）
                # -----------------------------
                yellow_reason = ""
                is_yellow_band = (confidence >= low_th and confidence < high_th)
                is_low_green_unstable = (confidence >= high_th and confidence < (high_th + 5))

                if is_yellow_band or (multifacet_tag and is_low_green_unstable):
                    yellow_reason = classify_yellow_reason_expected_vs_observed(
                        multifacet_tag=multifacet_tag,
                        facets_expected=facets_expected,
                        topics_observed_dist=topics_observed_dist,
                    )
                    # [print-YR1] ✅ 加在「下面」
                    if os.getenv("RAG_YELLOW_DEBUG", "0") == "1":
                        print(f"[print-YR1][下面] yellow_reason={yellow_reason} topic_dist={topic_dist} coverage_ratio={coverage_ratio}")

                    # 常態黃色原因輸出（不需 debug flag）
                    print(f"  [yellow_reason] {yellow_reason} (coverage={coverage_ratio})")

                if confidence >= high_th:
                    # 高信心 → 綠色
                    fill = FILL_GREEN
                    print(f"  ✅ 高信心題目，分數 {confidence:.1f} ≥ {high_th}，標記為綠色")

                elif confidence < low_th:
                    # 低信心 → 覆寫成未匹配（不顯示錯誤配對內容）+ 觸發 SKIP
                    print(f"  ⚠️  信心分數低於低信心閾值 ({confidence:.1f} < {low_th})，判定為未匹配（紅色）")
                    ws[f"{dept_col}{row_idx}"] = clean_excel_text("未匹配（請人工判定）")
                    ws[f"{impact_col}{row_idx}"] = clean_excel_text("知識庫無相關資料或語意不匹配，請人工填寫")
                    ws[f"{dept_col}{row_idx}"].alignment = ALIGN_D_COLUMN
                    ws[f"{impact_col}{row_idx}"].alignment = ALIGN_E_COLUMN
                    ws[f"{dept_col}{row_idx}"].fill = FILL_RED
                    ws[f"{impact_col}{row_idx}"].fill = FILL_RED
                    low_confidence_count += 1
                    results.append({
                        "row_index": row_idx,
                        "question": question,
                        "matched": False,
                        "confidence": confidence,
                        "department": "未匹配（請人工判定）",
                    })
                    # ✅ 補觸發 SKIP：低信心視為 NO（與 keyword_mismatch 路徑對齊）
                    if not no_skip_target:
                        no_skip_target = parse_no_skip_target(question)
                    if no_skip_target:
                        skip_until_clause = no_skip_target
                        skip_reason = f"{clause or row_idx} 低信心未匹配（視為 NO）→ 跳至 {no_skip_target}"
                        print(f"  ⚠️  低信心未匹配，啟動跳題：{skip_reason}")
                    continue  # 不計入 matched_count

                else:
                    # 中信心區間（55-70）→ 黃色，LLM 可升級
                    print(f"  ~ 信心分數介於 {low_th} 與 {high_th} 之間（黃色區間）")

                    if AZURE_OPENAI_ENABLED:
                        # 多面向題：合併所有 matches 的 dept/impact 傳給 LLM
                        if multifacet_tag and len(matches) > 1:
                            combined_impact = "\n".join(
                                m.get("impact", "") for m in matches if m.get("impact"))
                            combined_dept = "、".join(dict.fromkeys(
                                m.get("dept", "") for m in matches if m.get("dept")))
                            first_match = {
                                "dept": combined_dept,
                                "impact": combined_impact[:600],
                                "behavior": behavior,
                                "category": selected.get("category"),
                                "article": combined_impact[:400],
                                "score": confidence / 100.0,
                            }
                            print(f"  [多面向LLM] 合併 {len(matches)} 個 matches 傳給 LLM 審查")
                        else:
                            first_match = {
                                "dept": matches[0]["dept"],
                                "impact": matches[0]["impact"],
                                "behavior": behavior,
                                "category": selected.get("category"),
                                "article": selected.get("article"),
                                "score": matches[0]["score"],
                            }

                        llm_upgrade = azure_llm_upgrade_yellow(question, first_match)

                        if llm_upgrade is True:
                            fill = FILL_GREEN
                            print(f"  ✅ LLM 升級：黃色 → 綠色")
                        else:
                            fill = FILL_YELLOW
                            print(f"  ~ LLM 保持黃色（建議人工抽查）")
                            low_confidence_count += 1
                    else:
                        fill = FILL_YELLOW
                        print(f"  ~ 標記為黃色（建議人工抽查）")
                        low_confidence_count += 1

                if fill is not None:
                    ws[f"{dept_col}{row_idx}"].fill = fill
                    ws[f"{impact_col}{row_idx}"].fill = fill

                matched_count += 1

                # ========== ✅ 新增：匹配成功後也檢查 NO 跳題邏輯 ==========
                # 先解析題目是否有 NO 分支跳題指令
                if not no_skip_target:  # 避免重複解析
                    no_skip_target = parse_no_skip_target(question)

                # 如果題目有 NO 跳題指令，且答案是 NO，則啟動 skip
                if no_skip_target and is_answer_no(impact_text):
                    skip_until_clause = no_skip_target
                    skip_reason = f"{clause or row_idx} 答案為 NO → 跳至 {no_skip_target}"
                    # [print-SK3] ✅ 加在「下面」
                    if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
                        print(f"[print-SK3][下面] enable skip after match: {skip_reason}")
                    print(f"  ⚠️  偵測到 NO 答案，啟動跳題：{skip_reason}")

                results.append({
                    "row_index": row_idx,
                    "question": question,
                    "matched": True,
                    "confidence": confidence,
                    "department": department_text,
                    "status": impact_text[:100] + "..." if len(impact_text) > 100 else impact_text,
                    "behavior": behavior,
                    "match_count": len(matches),  # ⬅ 新增：匹配數量
                    "multifacet_tag": multifacet_tag,
                    "facets_expected": facets_expected,
                    "topics_observed_dist": topics_observed_dist,
                    "topic_dist": topic_dist,
                    "coverage_ratio": coverage_ratio,
                    "yellow_reason": yellow_reason,
                })

        else:
            print(f"  ✗ RAG 未找到匹配結果")
            ws[f"{dept_col}{row_idx}"] = "未匹配"
            ws[f"{impact_col}{row_idx}"] = "請人工填寫"

            # ★ D 欄置中，E 欄靠左（保持版面一致）
            ws[f"{dept_col}{row_idx}"].alignment = ALIGN_D_COLUMN
            ws[f"{impact_col}{row_idx}"].alignment = ALIGN_E_COLUMN

            # 標記為紅色
            ws[f"{dept_col}{row_idx}"].fill = FILL_RED
            ws[f"{impact_col}{row_idx}"].fill = FILL_RED

            low_confidence_count += 1

            results.append({
                "row_index": row_idx,
                "question": question,
                "matched": False,
                "confidence": 0,
                "department": "未匹配",
            })

            # ========== ✅ 新增：未匹配時也檢查 NO 跳題邏輯 ==========
            # 先解析題目是否有 NO 分支跳題指令（如果還沒解析過）
            if not no_skip_target:
                no_skip_target = parse_no_skip_target(question)

            # 如果題目有 NO 跳題指令，未匹配視為 NO，啟動 skip
            if no_skip_target:
                skip_until_clause = no_skip_target
                skip_reason = f"{clause or row_idx} 未匹配（視為 NO）→ 跳至 {no_skip_target}"
                # [print-SK4] ✅ 加在「下面」
                if os.getenv("RAG_SKIP_DEBUG", "0") == "1":
                    print(f"[print-SK4][下面] enable skip on no match: {skip_reason}")
                print(f"  ⚠️  未匹配且有 NO 跳題指令，啟動跳題：{skip_reason}")

    # ========== 4. 儲存結果 ==========
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"security_c_result_{timestamp}.xlsx"
    output_dir = os.path.dirname(survey_path)
    output_path = os.path.join(output_dir, output_filename)

    wb.save(output_path)
    wb.close()

    print(f"\n[儲存完成] 輸出檔案: {output_path}")

    # ========== 5. 產生報告 ==========
    total_questions = len(results)

    report = {
        "total_questions": total_questions,
        "matched_count": matched_count,
        "unmatched_count": total_questions - matched_count,
        "low_confidence_count": low_confidence_count,
        "match_rate": f"{matched_count / total_questions * 100:.1f}%" if total_questions > 0 else "0%",
        "details": results
    }

    print("\n" + "=" * 80)
    print("[處理摘要]")
    print(f"  總題數: {total_questions}")
    print(f"  成功匹配: {matched_count} ({report['match_rate']})")
    print(f"  未匹配: {total_questions - matched_count}")
    print(f"  低信心題目: {low_confidence_count}")
    print("=" * 80)

    return {
        "output_path": output_path,
        "report": report
    }


if __name__ == "__main__":
    """本地測試用"""
    import sys

    if len(sys.argv) < 3:
        print("使用方式: python rpa_security_c.py <問卷Excel> <資料庫Excel>")
        sys.exit(1)

    survey_file = sys.argv[1]
    db_file = sys.argv[2]

    result = process_workbook(survey_file, db_file, "測試客戶")
    print(f"\n✅ 完成！輸出檔案: {result['output_path']}")
