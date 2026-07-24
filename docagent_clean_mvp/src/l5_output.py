"""
L5 - 輸出層 (Output Layer)
============================
輸入：Question + Answer 清單
輸出：
  1. 回填後的 Excel（openpyxl 寫回答案欄位，依信心分數上色）
  2. JSON 稽核報告 (Audit Trail)，記錄每題的證據來源、分數、決策依據

設計原則 (Design Principles)
----------------------------
稽核報告是這個產品和「純 ChatGPT 問答」最大的差異化賣點——
每個答案都可以回溯到「哪份文件、哪一列/頁、相似度分數多少」，
這是企業合規稽核 (Compliance Audit) 願意付費的核心原因，不是黑盒子。
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .l2_schema import Question
from .l4_reasoning import Answer

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # 自動通過
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 建議覆核
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # 證據不足

STATUS_FILL = {
    "auto_approve": GREEN_FILL,
    "review_suggested": YELLOW_FILL,
    "insufficient_evidence": RED_FILL,
}


def write_excel(
    questionnaire_path: str,
    questions: List[Question],
    answers: Dict[str, Answer],
    output_path: str,
) -> None:
    wb = load_workbook(questionnaire_path)

    for q in questions:
        ans = answers.get(q.id)
        if ans is None:
            continue
        ws = wb[q.sheet]
        # 若原表沒有答案欄，就寫在問題欄右邊一格
        answer_col_idx = (q.answer_col if q.answer_col is not None else q.question_col + 1) + 1  # openpyxl 1-based
        row_idx = q.row + 1  # openpyxl 1-based

        cell = ws.cell(row=row_idx, column=answer_col_idx)
        cell.value = ans.answer_text
        cell.fill = STATUS_FILL.get(ans.status, RED_FILL)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_audit_json(
    questions: List[Question],
    answers: Dict[str, Answer],
    output_path: str,
) -> None:
    records = []
    for q in questions:
        ans = answers.get(q.id)
        if ans is None:
            continue
        records.append(
            {
                "question_id": q.id,
                "question_text": q.text,
                "sheet": q.sheet,
                "row": q.row,
                "answer_text": ans.answer_text,
                "confidence": ans.confidence,
                "citation": ans.citation,
                "status": ans.status,
                "needs_review": ans.needs_review,
            }
        )

    total = len(records)
    auto = sum(1 for r in records if r["status"] == "auto_approve")
    review = sum(1 for r in records if r["status"] == "review_suggested")
    insufficient = sum(1 for r in records if r["status"] == "insufficient_evidence")

    summary = {
        "total_questions": total,
        "auto_approved": auto,
        "review_suggested": review,
        "insufficient_evidence": insufficient,
        "automation_rate": round((auto + review) / total, 4) if total else 0.0,
        "auto_approve_rate": round(auto / total, 4) if total else 0.0,
    }

    payload = {"summary": summary, "records": records}
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    Path(output_path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
