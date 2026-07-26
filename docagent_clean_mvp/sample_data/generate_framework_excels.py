"""
產生 10 份合規框架比對用的示範 Excel（CAIQ / SIG / NIST / VSA / ISO 27001，各 2 份，每份 100 列）。

重要說明 (Legal Note)：CAIQ（Cloud Security Alliance）、SIG（Shared Assessments）、
VSA（Vendor Security Alliance）都是「有版權的正式問卷」，不能照抄原始題目文字。
這裡的欄位內容是根據這些框架公開已知的「控制領域」(Control Domain) 自己原創改寫的
範例敘述，只是拿框架名稱當標籤展示「引擎能對應多種框架」，不是真的複製那些機構的
問卷原文——正式對外demo/銷售時也要延用這個原則，避免版權爭議。
"""

from pathlib import Path
import random

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

OUT_DIR = Path(__file__).parent / "raw_excels"
OUT_DIR.mkdir(exist_ok=True)

random.seed(42)

HEADER = ["Index", "Domain", "Control ID", "Clause Description", "Owner", "Evidence Status"]

# 每個框架 2 份檔案共用的「領域池」，用來產生另外 99 列的填充資料（Filler Rows）
DOMAIN_POOL = {
    "CAIQ": ["Application & Interface Security", "Audit Assurance & Compliance", "Business Continuity Management",
             "Change Control & Configuration", "Data Security & Lifecycle", "Datacenter Security",
             "Encryption & Key Management", "Governance & Risk Management", "Human Resources Security",
             "Identity & Access Management", "Infrastructure & Virtualization", "Interoperability & Portability",
             "Mobile Security", "Security Incident Management", "Supply Chain Management", "Threat & Vulnerability Management"],
    "SIG": ["Risk Management", "Security Policy", "Asset & Information Management", "Human Resources Security",
            "Physical & Environmental Security", "IT Operations Management", "Access Control",
            "Application Security", "Incident Event Management", "Business Continuity Management",
            "Compliance", "Privacy", "Endpoint Security", "Network Security", "Cloud Hosting Services"],
    "NIST": ["Govern", "Identify", "Protect", "Detect", "Respond", "Recover"],
    "VSA": ["Data Protection", "Access Control", "Network Security", "Vulnerability Management",
            "Incident Response", "Business Continuity", "Third-Party Risk", "Physical Security",
            "Security Awareness Training", "Secure Development"],
    "ISO27001": ["Organizational Controls", "People Controls", "Physical Controls", "Technological Controls"],
}

OWNERS = ["IT Security", "Legal & Compliance", "IT Operations", "HR", "Facilities", "Procurement",
          "Quality Assurance", "Audit", "EHS", "Data Privacy Office"]

FILLER_VERBS = ["Review", "Audit", "Assessment", "Monitoring", "Training", "Certification Check",
                "Policy Update", "Control Testing", "Risk Scoring", "Remediation Tracking"]


def make_filler_row(idx: int, framework: str) -> list:
    domain = random.choice(DOMAIN_POOL[framework])
    verb = random.choice(FILLER_VERBS)
    control_id = f"{framework}-{idx:03d}"
    clause = f"Standard {domain} {verb.lower()} procedure is documented and reviewed on a recurring cycle. [{control_id}]"
    owner = random.choice(OWNERS)
    status = random.choice(["Compliant - evidence on file", "In progress - remediation scheduled",
                             "Compliant - last reviewed this cycle", "Pending - awaiting evidence upload"])
    return [idx, domain, control_id, clause, owner, status]


# 每個檔案指定一列「真正會被 demo 命中的答案列」，其餘 99 列用 filler 產生
# (framework, set_label, target_row, domain, control_id, clause, owner, status)
# 注意 (Consistency Note)：clause 統一用「Standard {Domain} control mandates that ...
# [control_id]」開頭+結尾的句式，跟 filler rows 的「Standard {domain} {verb}
# procedure is documented and reviewed on a recurring cycle. [control_id]」共用
# 同一個開頭字詞 "Standard"——這是第二輪修正：第一輪只統一了 Control ID 格式跟
# 後綴，但句子開頭沒統一（target 是 "Multi-factor authentication is..."，filler
# 是 "Standard XXX..."），使用者一眼掃過 Clause Description 欄位還是能靠「這句
# 沒有用 Standard 開頭」認出答案列。現在句首、句尾格式都跟 filler 同源，只有
# 中段的實際內容不同——這才是真正只靠「語意檢索」而非「格式」才能定位到答案。
TARGET_SPECS = [
    ("CAIQ", "SetA", 34, "Identity & Access Management", "CAIQ-034",
     "Standard Identity & Access Management control mandates that multi-factor authentication (MFA) is enforced for all privileged and remote access accounts. [CAIQ-034]",
     "IT Security", "Compliant - evidence on file"),
    ("CAIQ", "SetB", 61, "Encryption & Key Management", "CAIQ-061",
     "Standard Encryption & Key Management control mandates that all data at rest is encrypted using AES-256, with key rotation performed every 90 days. [CAIQ-061]",
     "IT Security", "Compliant - last reviewed this cycle"),
    ("SIG", "SetA", 18, "Access Control", "SIG-018",
     "Standard Access Control policy mandates that user access rights are reviewed and re-certified on a quarterly basis by data owners. [SIG-018]",
     "IT Security", "Compliant - evidence on file"),
    ("SIG", "SetB", 77, "Business Continuity Management", "SIG-077",
     "Standard Business Continuity Management control mandates that a documented Business Continuity Plan is tested via full failover simulation at least once per year. [SIG-077]",
     "IT Operations", "In progress - remediation scheduled"),
    ("NIST", "SetA", 9, "Protect", "NIST-009",
     "Standard Protect function control mandates that endpoint devices are protected by centrally managed anti-malware with automatic signature updates. [NIST-009]",
     "IT Security", "Compliant - last reviewed this cycle"),
    ("NIST", "SetB", 45, "Respond", "NIST-045",
     "Standard Respond function control mandates that a formal Incident Response Plan defines roles, escalation paths, and a 24-hour notification requirement for confirmed breaches. [NIST-045]",
     "IT Security", "Compliant - evidence on file"),
    ("VSA", "SetA", 52, "Third-Party Risk", "VSA-052",
     "Standard Third-Party Risk control mandates that subcontractors and fourth parties handling customer data are subject to the same security due-diligence review as direct vendors. [VSA-052]",
     "Procurement", "Compliant - evidence on file"),
    ("VSA", "SetB", 23, "Vulnerability Management", "VSA-023",
     "Standard Vulnerability Management control mandates that critical vulnerabilities identified by external scanning are remediated within 15 days of disclosure. [VSA-023]",
     "IT Security", "Pending - awaiting evidence upload"),
    ("ISO27001", "SetA", 88, "Technological Controls", "ISO27001-088",
     "Standard Technological Controls policy mandates that network segmentation isolates production, staging, and corporate environments with documented firewall rulesets. [ISO27001-088]",
     "IT Operations", "Compliant - last reviewed this cycle"),
    ("ISO27001", "SetB", 5, "Organizational Controls", "ISO27001-005",
     "Standard Organizational Controls policy mandates that an information security policy is approved by senior management and reviewed at least annually. [ISO27001-005]",
     "IT Security", "Compliant - evidence on file"),
]


def build_workbook(framework: str, set_label: str, target_row: int, domain: str,
                    control_id: str, clause: str, owner: str, status: str, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{framework}_{set_label}"

    ws.append(HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for idx in range(1, 101):
        if idx == target_row:
            ws.append([idx, domain, control_id, clause, owner, status])
        else:
            ws.append(make_filler_row(idx, framework))

    out_path = OUT_DIR / filename
    wb.save(out_path)
    print(f"已產生：{out_path.name}（目標列 {target_row}）")


def main():
    for framework, set_label, target_row, domain, control_id, clause, owner, status in TARGET_SPECS:
        filename = f"{framework}_{set_label}.xlsx"
        build_workbook(framework, set_label, target_row, domain, control_id, clause, owner, status, filename)


if __name__ == "__main__":
    main()
