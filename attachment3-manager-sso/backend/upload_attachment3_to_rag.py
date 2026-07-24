#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
上傳附件三到 RAG 知識庫
不動原本 server.py，直接操作資料庫插入 PdfFile + PdfChunk
每一列附件三 = 一個 chunk
"""

import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
import hashlib
import sys
import json
from datetime import datetime
from uuid import uuid4
from sqlmodel import Session, create_engine, SQLModel, Field, Column
from pgvector.sqlalchemy import Vector
from typing import Optional, List

# ========== LLM-first 萃取 ==========
import os
sys.path.insert(0, '/app')

# ✅ 對齊 rpa_security_c.py：優先使用新版函式
try:
    from llm_first_schema import (
        AzureOpenAIConfig,
        extract_knowledge_schema_llm,
    )
    _HAS_LLM_FIRST = True
except Exception as e:
    print(f"⚠️  LLM-first schema 模組載入失敗: {e}")
    _HAS_LLM_FIRST = False

# ========== 配置 ==========
# Azure OpenAI 配置（與 rpa_security_c.py 一致）
# ✅ 優先從環境變數讀取（避免把 key 寫死在程式）
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT", "https://en-openai01.openai.azure.com")
AZURE_OPENAI_API_KEY = os.getenv("AZURE_OPENAI_API_KEY", "1xPk95hy7EROtFp4yytQD6jrb237tGuewf0vGuMWExcbAr1TSgjUJQQJ99BFACHYHv6XJ3w3AAABACOGdXkj")
AZURE_OPENAI_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT", "gpt-5-chat")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview")

DB_URL = "postgresql://dgtk:dgtk@10.100.40.5:8002/dgtk"
EMBED_API = "http://10.100.40.5:8004/api/embed"
UNIT = "SYSTEM"
PDF_NAME = "附件三_EnvSafety_atta3_知識庫"
PROGRESS_FILE = "/app/attachment3_upload_progress.json"

# ========== 定義資料庫模型（與 server.py 一致） ==========
class PdfFile(SQLModel, table=True):
    id: Optional[str] = Field(primary_key=True, default_factory=lambda: str(uuid4()))
    hash: str
    unit: str
    name: str
    size: str
    date: Optional[datetime] = Field(default_factory=datetime.now)

class PdfChunk(SQLModel, table=True):
    id: Optional[str] = Field(primary_key=True, default_factory=lambda: str(uuid4()))
    pdf_id: str = Field(foreign_key="pdffile.id")
    page_hash: str
    chunk_index: int
    xywh: str
    text: str
    embed: List[float] = Field(sa_column=Column(Vector(1024)))


def update_progress(stage, percent, message=""):
    """更新進度到檔案"""
    progress = {
        "stage": stage,
        "percent": percent,
        "message": message,
        "timestamp": datetime.now().isoformat()
    }
    try:
        with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
            json.dump(progress, f, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️  無法寫入進度檔案: {e}", file=sys.stderr)


def main(attachment_path: str):
    print("=" * 80)
    print("上傳附件三到 RAG 知識庫（LLM-first 版本）")
    print("=" * 80)

    update_progress("init", 0, "開始處理")

    # ========== 1. 讀取附件三 ==========
    print(f"\n[步驟 1] 讀取附件三: {attachment_path}")
    update_progress("reading", 5, "讀取 Excel 檔案")

    wb = openpyxl.load_workbook(attachment_path, data_only=True)
    ws = wb.active

    print(f"  - 工作表名稱: {ws.title}")
    print(f"  - 最大行數: {ws.max_row}")

    # ========== 初始化 LLM 配置 ==========
    llm_cfg = None
    if _HAS_LLM_FIRST:
        try:
            llm_cfg = AzureOpenAIConfig(
                endpoint=AZURE_OPENAI_ENDPOINT,
                api_key=AZURE_OPENAI_API_KEY,
                deployment=AZURE_OPENAI_DEPLOYMENT,
                api_version=AZURE_OPENAI_API_VERSION,
            )
            print(f"  ✅ LLM-first 配置成功: {AZURE_OPENAI_ENDPOINT}/{AZURE_OPENAI_DEPLOYMENT}")
            # [print-A3CFG0][下面] LLM 設定確認（不印 key）
            print(f"  [print-A3CFG0][下面] endpoint={AZURE_OPENAI_ENDPOINT} deployment={AZURE_OPENAI_DEPLOYMENT} api_version={AZURE_OPENAI_API_VERSION}")
        except Exception as e:
            print(f"  ⚠️  LLM-first 配置失敗: {e}")
            llm_cfg = None
    else:
        print("  ⚠️  LLM-first 模組未載入，將使用傳統上傳")

    rows_data = []
    llm_success_count = 0
    llm_fallback_count = 0
    processed_count = 0
    valid_data_count = 0

    # ========== 步驟 1：收集所有需要處理的 row 資料 ==========
    print(f"  📊 開始收集資料...")
    row_tasks = []
    for row_idx in range(2, ws.max_row + 1):
        a_col = ws[f"A{row_idx}"].value or ""
        b_col = ws[f"B{row_idx}"].value or ""
        c_col = ws[f"C{row_idx}"].value or ""
        d_col = ws[f"D{row_idx}"].value or ""
        e_col = ws[f"E{row_idx}"].value or ""
        f_col = ws[f"F{row_idx}"].value or ""
        g_col = ws[f"G{row_idx}"].value or ""
        
        if not any([a_col, b_col, c_col, d_col, e_col, f_col]):
            continue
        
        row_id = f"{ws.title}!A{row_idx}"
        row_tasks.append((row_idx, row_id, a_col, b_col, c_col, d_col, e_col, f_col, g_col))

    print(f"  ✅ 收集完成，共 {len(row_tasks)} 筆待處理")

    # ========== 步驟 2：定義單個 row 的處理函數 ==========
    def process_single_row(task_data):
        """並行處理單個 row 的 LLM 萃取"""
        row_idx, row_id, a_col, b_col, c_col, d_col, e_col, f_col, g_col = task_data
        
        knowledge_item = None
        success = False
        
        if llm_cfg:
            try:
                keywords_list = [kw.strip() for kw in str(c_col).split("\n") if kw.strip()]
                knowledge_item = extract_knowledge_schema_llm(
                    cfg=llm_cfg,
                    row_id=row_id,
                    category=str(a_col),
                    behavior=str(b_col),
                    keywords=keywords_list,
                    article=str(d_col),
                    dept=str(e_col),
                    impact=str(f_col),
                    source=str(g_col),
                    debug=False
                )
                
                if getattr(knowledge_item, "confidence", 0.0) > 0.0:
                    success = True
            except Exception as e:
                print(f"  ⚠️  row {row_id} LLM 萃取失敗: {e}")
        
        # Fallback
        if knowledge_item is None:
            keywords_list = [kw.strip() for kw in str(c_col).split("\n") if kw.strip()]
            knowledge_item = type('obj', (object,), {
                'row_id': row_id,
                'category': str(a_col),
                'behavior': str(b_col),
                'intent': "",
                'topic_tags': [],
                'keywords': keywords_list,
                'article': str(d_col),
                'dept': str(e_col),
                'impact': str(f_col),
                'source': str(g_col),
                'confidence': 0.0
            })()
        
        return (row_idx, row_id, a_col, b_col, c_col, d_col, e_col, f_col, g_col, knowledge_item, success)

    # ========== 步驟 3：並行處理（5 個 worker）==========
    print(f"  🚀 開始並行 LLM 萃取（5 個並發）...")
    results = []
    
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_task = {executor.submit(process_single_row, task): task for task in row_tasks}
        
        for future in as_completed(future_to_task):
            try:
                result = future.result()
                results.append(result)
                
                row_idx, row_id, a_col, b_col, c_col, d_col, e_col, f_col, g_col, knowledge_item, success = result
                
                if success:
                    llm_success_count += 1
                else:
                    llm_fallback_count += 1
                
                valid_data_count += 1
                
                # 更新進度（每 10 筆）
                if valid_data_count % 10 == 0:
                    estimated_total = len(row_tasks)
                    progress_percent = 5 + min(int((valid_data_count / estimated_total) * 5), 5)
                    update_progress("reading", progress_percent, f"LLM 萃取中：{valid_data_count}/~{estimated_total} 筆 ({int((valid_data_count/estimated_total)*100) if estimated_total > 0 else 0}%)")
                
                # 抽樣顯示前 3 筆
                if valid_data_count <= 3:
                    print(f"  [LLM-first] row={row_id}")
                    print(f"    intent: {str(getattr(knowledge_item, 'intent', ''))[:50]}...")
                    print(f"    tags: {(getattr(knowledge_item, 'topic_tags', []) or [])[:3]}")
                    print(f"    confidence: {float(getattr(knowledge_item, 'confidence', 0.0)):.2f}")
            
            except Exception as e:
                print(f"  ❌ 處理失敗: {e}")

    # ========== 步驟 4：按照原始順序排序並組裝 chunks ==========
    results.sort(key=lambda x: x[0])  # 按 row_idx 排序
    
    for result in results:
        row_idx, row_id, a_col, b_col, c_col, d_col, e_col, f_col, g_col, knowledge_item, success = result
        
        # 組成結構化文字
        try:
            if hasattr(knowledge_item, 'model_dump'):
                schema_json = json.dumps(knowledge_item.model_dump(), ensure_ascii=False)
            else:
                schema_json = json.dumps({
                    'row_id': getattr(knowledge_item, 'row_id', row_id),
                    'category': getattr(knowledge_item, 'category', ''),
                    'behavior': getattr(knowledge_item, 'behavior', ''),
                    'intent': getattr(knowledge_item, 'intent', ''),
                    'topic_tags': getattr(knowledge_item, 'topic_tags', []),
                    'keywords': getattr(knowledge_item, 'keywords', []),
                    'article': getattr(knowledge_item, 'article', ''),
                    'dept': getattr(knowledge_item, 'dept', ''),
                    'impact': getattr(knowledge_item, 'impact', ''),
                    'source': getattr(knowledge_item, 'source', ''),
                    'confidence': getattr(knowledge_item, 'confidence', 0.0),
                }, ensure_ascii=False)
        except Exception:
            schema_json = "{}"

        # ✅ 方案 B：核心語意置頂重複（提高 embedding 語意權重）
        intent = getattr(knowledge_item, 'intent', '')
        article = getattr(knowledge_item, 'article', str(d_col))
        topic_tags_str = ', '.join(getattr(knowledge_item, 'topic_tags', []) or [])
        
        chunk_text = f"""語意核心：{intent}

條文核心：{article}

主題：{topic_tags_str}

[schema_json] {schema_json}
[row_id] {getattr(knowledge_item, 'row_id', row_id)}
[類別] {getattr(knowledge_item, 'category', str(a_col))}
[行為準則] {getattr(knowledge_item, 'behavior', str(b_col))}
[語意摘要] {intent}
[主題標籤] {topic_tags_str}
[關鍵字] {', '.join(getattr(knowledge_item, 'keywords', []) or [])}
[條文內容] {article}
[權責部門] {getattr(knowledge_item, 'dept', str(e_col))}
[影響] {getattr(knowledge_item, 'impact', str(f_col))}
[出處] {getattr(knowledge_item, 'source', str(g_col))}
"""
        
        rows_data.append({
            "row_id": row_id,
            "text": chunk_text,
            "knowledge_item": knowledge_item
        })

    wb.close()

    print(f"\n  ✅ 讀取完成，共 {len(rows_data)} 列有效資料")
    if rows_data and llm_cfg:
        print(f"  📊 LLM 萃取統計:")
        print(f"     - 成功萃取: {llm_success_count} 筆 ({llm_success_count/len(rows_data)*100:.1f}%)")
        print(f"     - Fallback: {llm_fallback_count} 筆 ({llm_fallback_count/len(rows_data)*100:.1f}%)")

    update_progress("reading", 10, f"讀取完成（LLM 萃取 {llm_success_count}/{len(rows_data)} 筆）")

    if not rows_data:
        print("  ❌ 沒有資料可上傳，結束")
        update_progress("error", 0, "沒有資料可上傳")
        return

    # ========== 2. 呼叫 embedding API（分批處理） ==========
    print(f"\n[步驟 2] 呼叫 Embedding API: {EMBED_API}")
    update_progress("embedding", 15, "開始生成向量")

    texts = [row["text"] for row in rows_data]
    vectors = []
    batch_size = 15

    try:
        total_batches = (len(texts) + batch_size - 1) // batch_size
        print(f"  📝 總共 {len(texts)} 筆資料，分成 {total_batches} 批處理...")

        for i in range(0, len(texts), batch_size):
            batch_texts = texts[i:i + batch_size]
            batch_num = i // batch_size + 1

            print(f"  📡 處理第 {batch_num}/{total_batches} 批（{len(batch_texts)} 筆）...")

            progress_percent = 15 + int((batch_num / total_batches) * 35)
            update_progress("embedding", progress_percent, f"生成向量 {batch_num}/{total_batches} 批")

            resp = requests.post(EMBED_API, json={"input": batch_texts}, timeout=120)
            resp.raise_for_status()
            result = resp.json()

            if "vectors" in result:
                batch_vectors = result["vectors"]
            elif "data" in result:
                batch_vectors = [item["embedding"] for item in result["data"]]
            else:
                raise ValueError(f"無法解析 embedding 回傳格式: {list(result.keys())}")

            vectors.extend(batch_vectors)
            print(f"     ✅ 第 {batch_num} 批完成（累計 {len(vectors)}/{len(texts)}）")

        print(f"\n  ✅ 全部 Embedding 完成")
        print(f"     - 向量數量: {len(vectors)}")
        print(f"     - 向量維度: {len(vectors[0]) if vectors else 0}")

    except Exception as e:
        print(f"  ❌ Embedding 失敗: {e}")
        update_progress("error", 0, f"Embedding 失敗: {str(e)}")
        return

    # ========== 3. 插入資料庫 ==========
    print(f"\n[步驟 3] 插入資料庫: {DB_URL}")
    update_progress("database", 50, "開始寫入資料庫")

    engine = create_engine(DB_URL, echo=False)

    with Session(engine) as db:
        existing = db.query(PdfFile).filter(PdfFile.name == PDF_NAME, PdfFile.unit == UNIT).first()

        if existing:
            print(f"  ⚠️  發現已存在的檔案，將先刪除舊資料...")
            update_progress("database", 55, "刪除舊資料")
            db.query(PdfChunk).filter(PdfChunk.pdf_id == existing.id).delete()
            db.delete(existing)
            db.commit()
            print(f"  ✅ 已刪除舊資料")

        update_progress("database", 60, "建立檔案記錄")
        pdf_file_id = str(uuid4())
        pdf_hash = hashlib.md5(PDF_NAME.encode()).hexdigest()

        pdf_file = PdfFile(
            id=pdf_file_id,
            hash=pdf_hash,
            unit=UNIT,
            name=PDF_NAME,
            size=str(len(rows_data)),
            date=datetime.now()
        )
        db.add(pdf_file)
        db.commit()

        print(f"  ✅ 建立 PdfFile:")
        print(f"     - pdf_id: {pdf_file_id}")
        print(f"     - name: {PDF_NAME}")

        print(f"  📝 建立 {len(rows_data)} 個 chunks...")
        update_progress("database", 65, "開始寫入資料")

        for idx, (row_data, vector) in enumerate(zip(rows_data, vectors)):
            chunk = PdfChunk(
                id=str(uuid4()),
                pdf_id=pdf_file_id,
                page_hash=row_data['row_id'],
                chunk_index=idx,
                xywh="",
                text=row_data["text"],
                embed=vector
            )
            db.add(chunk)

            # ✅ 每 10 筆更新一次進度（65% → 95%）
            if (idx + 1) % 10 == 0 or (idx + 1) == len(rows_data):
                progress = 65 + int(((idx + 1) / len(rows_data)) * 30)
                update_progress("database", progress, f"寫入資料 {idx + 1}/{len(rows_data)} 筆")


            if (idx + 1) % 50 == 0:
                print(f"     - 進度: {idx + 1}/{len(rows_data)}")

        update_progress("database", 95, "提交資料庫")
        db.commit()
        print(f"  ✅ 資料庫提交完成！")

    # ========== 4. 完成報告 ==========
    update_progress("complete", 100, "上傳完成")
    print("\n" + "=" * 80)
    print("✅ 上傳完成！")
    print("=" * 80)
    print(f"  PDF ID: {pdf_file_id}")
    print(f"  檔案名稱: {PDF_NAME}")
    print(f"  資料筆數: {len(rows_data)}")
    print("=" * 80)

    # 輸出 JSON 格式結果
    print("\n=== JSON_OUTPUT_START ===")
    print(json.dumps({
        "success": True,
        "pdf_id": pdf_file_id,
        "pdf_name": PDF_NAME,
        "chunk_count": len(rows_data),
        "unit": UNIT,
        "llm_success": llm_success_count,
        "llm_fallback": llm_fallback_count
    }, ensure_ascii=False))
    print("=== JSON_OUTPUT_END ===")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("❌ 錯誤：缺少檔案路徑參數")
        print("用法: python upload_attachment3_to_rag.py <附件三檔案路徑>")
        sys.exit(1)

    attachment_path = sys.argv[1]

    if not os.path.exists(attachment_path):
        print(f"❌ 錯誤：檔案不存在 - {attachment_path}")
        update_progress("error", 0, f"檔案不存在: {attachment_path}")
        sys.exit(1)

    try:
        main(attachment_path)
    except KeyboardInterrupt:
        print("\n\n⚠️  使用者中斷")
        update_progress("error", 0, "使用者中斷")
    except Exception as e:
        print(f"\n\n❌ 錯誤: {e}")
        update_progress("error", 0, f"錯誤: {str(e)}")
